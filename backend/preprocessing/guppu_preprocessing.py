import os
import shutil
import pandas as pd
import calendar
import re
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import firebase_admin
from firebase_admin import credentials, storage
from backend.utils.secrets_manager import get_firebase_secret

class GuppuPreprocessor:
    def __init__(self):
        self.download_dir = str(Path.home() / "Downloads")
        self.setup_firebase()
    
    def setup_firebase(self):
        """Firebase Storage 연결 설정"""
        try:
            from dotenv import load_dotenv
            
            load_dotenv()
            cred_dict = get_firebase_secret()
            
            BUCKET_NAME = os.getenv("STORAGE_BUCKET", "services-e42af.firebasestorage.app")
            
            # 기존 앱이 있으면 삭제하고 새로 초기화
            if firebase_admin._apps:
                firebase_admin.delete_app(firebase_admin.get_app())
            
            firebase_admin.initialize_app(
                credentials.Certificate(cred_dict),
                {"storageBucket": BUCKET_NAME}
            )
            
            self.bucket = storage.bucket()
            print("Firebase Storage 연결 완료")
        except Exception as e:
            print(f"Firebase 연결 실패: {e}")
            self.bucket = None
    
    def download_guppu_template(self):
        """Firebase에서 guppeu.xlsx 템플릿 다운로드"""
        if not self.bucket:
            print("Firebase 연결이 없습니다")
            return None
        
        try:
            blob = self.bucket.blob("guppeu.xlsx")
            temp_dir = os.path.join(os.getcwd(), "temp_processing")
            os.makedirs(temp_dir, exist_ok=True)
            
            local_path = os.path.join(temp_dir, "guppeu.xlsx")
            blob.download_to_filename(local_path)
            print(f"guppeu.xlsx 템플릿 다운로드 완료: {local_path}")
            return local_path
        except Exception as e:
            print(f"guppeu.xlsx 다운로드 실패: {e}")
            return None
    
    def get_bill_amount(self, company_name="구쁘"):
        """고지서에서 업데이트된 금액 조회"""
        try:
            import requests
            response = requests.get('http://localhost:5001/api/bill-amounts')
            if response.ok:
                bill_data = response.json()
                if company_name in bill_data:
                    amount_str = bill_data[company_name].get('amount', '')
                    if amount_str:
                        amount_clean = amount_str.replace(',', '').replace('원', '').strip()
                        if amount_clean.isdigit():
                            amount = float(amount_clean)
                            print(f"{company_name} 고지서 금액 조회: {amount:,.0f}원")
                            return amount
                        else:
                            print(f"금액 형식 오류: {amount_str}")
                            return None
                else:
                    print(f"{company_name} 고지서 정보를 찾을 수 없습니다")
                    return None
            else:
                print("고지서 정보 조회 실패")
                return None
        except Exception as e:
            print(f"고지서 금액 조회 실패: {e}")
            return None
    
    def find_uploaded_sms_file(self, company_name="구쁘"):
        """업로드된 구쁘 SMS 발송이력 파일 찾기"""
        try:
            temp_dir = os.path.join(os.getcwd(), "temp_processing")
            if not os.path.exists(temp_dir):
                print("temp_processing 폴더가 없습니다")
                return None
            
            files = os.listdir(temp_dir)
            for filename in files:
                if company_name in filename and ("SMS" in filename or "발송이력" in filename) and filename.endswith(('.xlsx', '.xls')):
                    file_path = os.path.join(temp_dir, filename)
                    print(f"구쁘 SMS 발송이력 파일 발견: {filename}")
                    return file_path
            
            print("구쁘 SMS 발송이력 파일을 찾을 수 없습니다")
            return None
        except Exception as e:
            print(f"SMS 파일 찾기 실패: {e}")
            return None
    
    def analyze_sms_data(self, sms_file_path):
        """SMS 발송이력 데이터 분석하여 문자 유형별 카운트"""
        try:
            print(f"SMS 발송이력 분석 시작: {sms_file_path}")
            
            # Excel 파일 읽기
            df = pd.read_excel(sms_file_path)
            
            if df.empty:
                print("SMS 데이터가 비어있습니다")
                return {"SMS": 0, "LMS": 0, "MMS": 0, "TALK": 0}
            
            print(f"총 {len(df)}개의 SMS 발송 기록 발견")
            print(f"컬럼 정보: {list(df.columns)}")
            
            # F열(제목), H열(상태), I열(문자유형) 확인
            title_col = None
            status_col = None
            msg_type_col = None
            
            # 컬럼 인덱스로 접근 (F열=5, H열=7, I열=8)
            if len(df.columns) > 5:
                title_col = df.columns[5]  # F열
                print(f"제목 컬럼 (F열): {title_col}")
            
            if len(df.columns) > 7:
                status_col = df.columns[7]  # H열
                print(f"상태 컬럼 (H열): {status_col}")
            
            if len(df.columns) > 8:
                msg_type_col = df.columns[8]  # I열
                print(f"문자유형 컬럼 (I열): {msg_type_col}")
            
            if title_col is None or status_col is None or msg_type_col is None:
                print("필요한 컬럼을 찾을 수 없습니다")
                return {"SMS": 0, "LMS": 0, "MMS": 0, "TALK": 0}
            
            # 카운트 초기화
            counts = {"SMS": 0, "LMS": 0, "MMS": 0, "TALK": 0}
            
            for idx, row in df.iterrows():
                title = str(row[title_col]) if pd.notna(row[title_col]) else ""
                status = str(row[status_col]) if pd.notna(row[status_col]) else ""
                msg_type = str(row[msg_type_col]) if pd.notna(row[msg_type_col]) else ""
                
                # 첫 번째 조건: H열(상태)이 "성공"인 것만 카운트
                if status != "성공":
                    continue
                
                # 제목에 "SMS_"가 포함된 경우 무조건 SMS로 분류
                if "SMS_" in title:
                    counts["SMS"] += 1
                elif "LMS/MMS" in msg_type:
                    counts["LMS"] += 1
                elif "TALK(알림톡)" in msg_type or "알림톡" in msg_type:
                    counts["TALK"] += 1
                # MMS는 pass (카운트하지 않음)
            
            print(f"SMS 분석 결과 (상태='성공'인 것만 카운트):")
            print(f"  - SMS: {counts['SMS']}개 (제목에 SMS_ 포함)")
            print(f"  - LMS: {counts['LMS']}개 (문자유형이 LMS/MMS)")
            print(f"  - MMS: {counts['MMS']}개 (pass)")
            print(f"  - 알림톡: {counts['TALK']}개 (문자유형이 TALK(알림톡))")
            
            return counts
            
        except Exception as e:
            print(f"SMS 데이터 분석 실패: {e}")
            return {"SMS": 0, "LMS": 0, "MMS": 0, "TALK": 0}

    def calculate_ics_usage_data(self, collection_date):
        """ICS 사용 계정 현황 데이터 계산"""
        try:
            collection_date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            year = collection_date_obj.year
            month = collection_date_obj.month
            
            # 해당 월의 마지막 일수 계산
            last_day = calendar.monthrange(year, month)[1]
            
            print(f"수집 날짜: {collection_date} ({year}년 {month}월)")
            print(f"해당 월의 마지막 일수: {last_day}일")
            
            # ICS 사용 현황 데이터 생성
            ics_data = {
                "last_day": last_day,  # E35 셀에 들어갈 값 (숫자)
                "days_30": 1 if last_day >= 30 else 0,  # AJ37-38 (30일) - 숫자
                "days_31": 1 if last_day >= 31 else 0,  # AK37-38 (31일) - 숫자
                "is_february": month == 2  # 2월 특별 처리용
            }
            
            print(f"ICS 데이터:")
            print(f"  - 마지막 일수 (E35): {ics_data['last_day']}일")
            print(f"  - 30일 사용 (AJ37-38): {ics_data['days_30']}")
            print(f"  - 31일 사용 (AK37-38): {ics_data['days_31']}")
            print(f"  - 2월 여부: {ics_data['is_february']}")
            
            return ics_data
            
        except Exception as e:
            print(f"ICS 사용 데이터 계산 실패: {e}")
            return {
                "last_day": 31,
                "days_30": 1, 
                "days_31": 1,
                "is_february": False
            }

    def calculate_amount_without_vat(self, total_amount):
        """부가세 10%를 제외한 금액 계산 (1원 오차 보정 포함)"""
        try:
            # 부가세 포함 금액에서 부가세 제외
            # 총금액 = 공급가액 + 부가세(공급가액의 10%)
            # 총금액 = 공급가액 × 1.1
            # 공급가액 = 총금액 / 1.1
            amount_without_vat = round(total_amount / 1.1)
            
            # 검증: 부가세 제외 금액 + 부가세 = 실제 고지서 금액인지 확인
            calculated_total = round(amount_without_vat * 1.1)
            difference = total_amount - calculated_total
            
            if difference != 0:
                # 1원 차이가 있으면 부가세 제외 금액을 보정
                amount_without_vat += difference
                print(f"1원 오차 보정: {difference:+d}원 조정")
                
            # 최종 검증
            final_total = round(amount_without_vat * 1.1)
            print(f"부가세 제외 계산: {total_amount:,}원 → {amount_without_vat:,}원 (검증: {final_total:,}원)")
            
            return amount_without_vat
        except Exception as e:
            print(f"부가세 제외 계산 실패: {e}")
            return None
    
    def update_guppu_template(self, template_path, amount_without_vat, collection_date, total_amount, sms_counts=None, ics_data=None):
        """구쁘 템플릿 업데이트"""
        try:
            workbook = load_workbook(template_path)
            
            # 수집 날짜로부터 년월 정보 생성
            collection_date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            year_month = f"{collection_date_obj.year}년 {collection_date_obj.month:02d}월"
            
            # 1. 시트명 변경 (예: "2025년 08월" → 수집한 달로 변경)
            for sheet in workbook.worksheets:
                if re.match(r'\d{4}년 \d{1,2}월', sheet.title):
                    old_title = sheet.title
                    sheet.title = year_month
                    print(f"시트명 변경: {old_title} → {year_month}")
                    break
            
            # 2. 통신서비스 이용료 시트 업데이트
            if '통신서비스 이용료' in workbook.sheetnames:
                main_sheet = workbook['통신서비스 이용료']
                
                # D2 셀에 기준월 입력 (YYYY-MM 형식)
                year_month_format = collection_date_obj.strftime('%Y-%m')
                main_sheet.cell(row=2, column=4).value = year_month_format
                print(f"통신서비스 이용료 시트 D2 셀 업데이트: {year_month_format}")
            
            # 3. 세부내역 시트 업데이트
            if '세부내역' in workbook.sheetnames:
                detail_sheet = workbook['세부내역']
                
                # E4 셀에 부가세 제외 금액 입력
                detail_sheet.cell(row=4, column=5).value = amount_without_vat
                print(f"세부내역 시트 E4 셀 업데이트: {amount_without_vat:,}원")
                
                # E7 셀에 부가세 10% 계산해서 입력
                vat_amount = round(amount_without_vat * 0.1)
                detail_sheet.cell(row=7, column=5).value = vat_amount
                print(f"세부내역 시트 E7 셀 업데이트: {vat_amount:,}원 (부가세 10%)")
                
                # E8 셀에 총 금액 입력 (고지서 금액과 동일하게)
                detail_sheet.cell(row=8, column=5).value = total_amount
                print(f"세부내역 시트 E8 셀 업데이트: {total_amount:,}원 (총 금액)")
                
                # SMS 카운트 데이터가 있는 경우 D12-D15에 입력
                if sms_counts:
                    # D12: SMS 카운트
                    detail_sheet.cell(row=12, column=4).value = sms_counts["SMS"]
                    print(f"세부내역 시트 D12 셀 업데이트 (SMS): {sms_counts['SMS']}개")
                    
                    # D13: LMS 카운트
                    detail_sheet.cell(row=13, column=4).value = sms_counts["LMS"]
                    print(f"세부내역 시트 D13 셀 업데이트 (LMS): {sms_counts['LMS']}개")
                    
                    # D14: MMS 카운트 (pass - 0으로 설정)
                    detail_sheet.cell(row=14, column=4).value = sms_counts["MMS"]
                    print(f"세부내역 시트 D14 셀 업데이트 (MMS): {sms_counts['MMS']}개")
                    
                    # D15: 알림톡 카운트
                    detail_sheet.cell(row=15, column=4).value = sms_counts["TALK"]
                    print(f"세부내역 시트 D15 셀 업데이트 (알림톡): {sms_counts['TALK']}개")
                
                # ICS 사용 계정 현황 데이터가 있는 경우 처리
                if ics_data:
                    # E35: 해당 월의 마지막 일수
                    detail_sheet.cell(row=35, column=5).value = ics_data["last_day"]
                    print(f"세부내역 시트 E35 셀 업데이트 (월 마지막 일수): {ics_data['last_day']}일")
                    
                    # AJ37-38 (30일): 30일까지 있는 월이면 "1", 아니면 "0"
                    aj_col = 36  # AJ열 = 36번째 컬럼
                    detail_sheet.cell(row=37, column=aj_col).value = ics_data["days_30"]  # 인바운드
                    detail_sheet.cell(row=38, column=aj_col).value = ics_data["days_30"]  # 아웃바운드
                    print(f"세부내역 시트 AJ37-38 셀 업데이트 (30일): {ics_data['days_30']}")
                    
                    # AK37-38 (31일): 31일까지 있는 월이면 "1", 아니면 "0"
                    ak_col = 37  # AK열 = 37번째 컬럼
                    detail_sheet.cell(row=37, column=ak_col).value = ics_data["days_31"]  # 인바운드
                    detail_sheet.cell(row=38, column=ak_col).value = ics_data["days_31"]  # 아웃바운드
                    print(f"세부내역 시트 AK37-38 셀 업데이트 (31일): {ics_data['days_31']}")
                    
                    # 2월 특별 처리: AH37-38 (28일)까지만 "1", 나머지는 "0"
                    if ics_data["is_february"]:
                        ah_col = 34  # AH열 = 34번째 컬럼
                        ai_col = 35  # AI열 = 35번째 컬럼
                        
                        # AH37-38 (28일)은 1
                        detail_sheet.cell(row=37, column=ah_col).value = 1  # 인바운드
                        detail_sheet.cell(row=38, column=ah_col).value = 1  # 아웃바운드
                        print(f"세부내역 시트 AH37-38 셀 업데이트 (28일): 1")
                        
                        # AI37-38 (29일)부터 AK37-38 (31일)까지는 0
                        for col in [ai_col, aj_col, ak_col]:  # AI, AJ, AK
                            detail_sheet.cell(row=37, column=col).value = 0  # 인바운드
                            detail_sheet.cell(row=38, column=col).value = 0  # 아웃바운드
                        print(f"세부내역 시트 AI37-38, AJ37-38, AK37-38 셀 업데이트 (29-31일): 0")
            
            # 4. 대외공문 시트 업데이트 (SK일렉링크와 동일한 로직)
            if '대외공문' in workbook.sheetnames:
                doc_sheet = workbook['대외공문']
                
                # B13 셀 업데이트: "제       목 : 2025년 07월 상담솔루션 서비스 수수료 정산 요청"
                b13_cell = doc_sheet.cell(row=13, column=2)
                if b13_cell.value and isinstance(b13_cell.value, str):
                    old_text = b13_cell.value
                    new_text = re.sub(r'\d{4}년 \d{1,2}월', year_month, old_text)
                    b13_cell.value = new_text
                    print(f"대외공문 B13 셀 업데이트: {old_text} → {new_text}")
                
                # B16 셀 업데이트 (B,C,D,E,F,G 16행 병합)
                b16_cell = doc_sheet.cell(row=16, column=2)
                if b16_cell.value and isinstance(b16_cell.value, str):
                    old_text = b16_cell.value
                    new_text = re.sub(r'2025년\d{1,2}월', year_month, old_text)
                    if new_text == old_text:
                        new_text = re.sub(r'\d{4}년\d{1,2}월', year_month, old_text)
                    b16_cell.value = new_text
                    print(f"대외공문 B16 셀 업데이트: {old_text} → {new_text}")
                
                # 하단 테이블 수식 업데이트 (B24, D24, D25)
                self.update_formula_references(doc_sheet, year_month)
            
            # 파일 저장 (YYMM 형식으로)
            collection_date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            date_prefix = f"{str(collection_date_obj.year)[2:]}{collection_date_obj.month:02d}"
            final_filename = f"{date_prefix}_구쁘_상담솔루션 청구내역서.xlsx"
            final_path = os.path.join(self.download_dir, final_filename)
            workbook.save(final_path)
            workbook.close()
            
            print(f"구쁘 청구내역서 생성 완료: {final_filename}")
            return final_path
            
        except Exception as e:
            print(f"템플릿 업데이트 실패: {e}")
            return None
    
    def update_formula_references(self, doc_sheet, year_month):
        """대외공문 시트의 수식에서 시트명 참조 업데이트"""
        try:
            # 수식이 있을 수 있는 셀들 확인 (24행, 25행 등)
            cells_to_check = [
                (24, 2), (24, 3), (24, 4),  # B24, C24, D24
                (25, 2), (25, 3), (25, 4),  # B25, C25, D25
            ]
            
            for row, col in cells_to_check:
                cell = doc_sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    old_formula = cell.value
                    # 수식에서 시트명 참조 패턴 찾아서 교체
                    new_formula = re.sub(r"'(\d{4}년 \d{1,2}월)'!", f"'{year_month}'!", old_formula)
                    if new_formula != old_formula:
                        cell.value = new_formula
                        print(f"대외공문 {chr(64+col)}{row} 셀 수식 업데이트: {old_formula} → {new_formula}")
                        
        except Exception as e:
            print(f"수식 참조 업데이트 오류: {e}")
    
    def process_guppu_data(self, collection_date):
        """구쁘 데이터 전처리 메인 함수"""
        try:
            print("=== 구쁘 데이터 전처리 시작 ===")
            
            # 1. Firebase에서 템플릿 다운로드
            template_path = self.download_guppu_template()
            if not template_path:
                print("템플릿 다운로드 실패")
                return False
            
            # 2. 고지서 금액 조회
            total_amount = self.get_bill_amount("구쁘")
            if total_amount is None:
                print("고지서 금액 조회 실패")
                return False
            
            # 3. 부가세 제외 금액 계산
            amount_without_vat = self.calculate_amount_without_vat(total_amount)
            if amount_without_vat is None:
                print("부가세 제외 계산 실패")
                return False
            
            # 4. SMS 발송이력 파일 찾기 및 분석
            sms_counts = {"SMS": 0, "LMS": 0, "MMS": 0, "TALK": 0}
            sms_file_path = self.find_uploaded_sms_file("구쁘")
            if sms_file_path:
                print("SMS 발송이력 파일 발견, 분석 시작...")
                sms_counts = self.analyze_sms_data(sms_file_path)
            else:
                print("SMS 발송이력 파일을 찾을 수 없어 기본값(0)으로 설정합니다")
            
            # 5. ICS 사용 계정 현황 데이터 계산
            ics_data = self.calculate_ics_usage_data(collection_date)
            
            # 6. 템플릿 업데이트 (SMS 카운트 + ICS 데이터 포함)
            final_invoice_path = self.update_guppu_template(template_path, amount_without_vat, collection_date, total_amount, sms_counts, ics_data)
            if not final_invoice_path:
                print("템플릿 업데이트 실패")
                return False
            
            print("=== 구쁘 데이터 전처리 완료 ===")
            return True
            
        except Exception as e:
            print(f"구쁘 데이터 전처리 실패: {e}")
            return False
