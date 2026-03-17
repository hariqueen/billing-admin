import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl.styles import PatternFill, Border, Side
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import re
import shutil
import firebase_admin
from firebase_admin import credentials, storage, firestore
from backend.utils.secrets_manager import get_firebase_secret
import json
import calendar

class KolonPreprocessor:
    def __init__(self):
        self.download_dir = os.path.join(os.getcwd(), "temp_processing")
        os.makedirs(self.download_dir, exist_ok=True)
        self.bucket = None
        self.db = None
        self.setup_firebase()
    
    def setup_firebase(self):
        """Firebase Storage 연결 설정"""
        try:
            cred_dict = get_firebase_secret()
            BUCKET_NAME = "services-e42af.firebasestorage.app"
            
            # 기존 앱이 있으면 삭제하고 새로 초기화
            if firebase_admin._apps:
                firebase_admin.delete_app(firebase_admin.get_app())
            
            firebase_admin.initialize_app(
                credentials.Certificate(cred_dict),
                {"storageBucket": BUCKET_NAME}
            )
            
            self.bucket = storage.bucket()
            self.db = firestore.client()
            print("Firebase Storage 및 Firestore 연결 완료")
        except Exception as e:
            print(f"Firebase 연결 실패: {e}")
            self.bucket = None
            self.db = None
    
    def download_kolon_template(self):
        """Firebase에서 kolon.xlsx 템플릿 다운로드"""
        if not self.bucket:
            print("Firebase 연결이 없습니다")
            return None
        
        try:
            blob = self.bucket.blob("kolon.xlsx")
            temp_dir = os.path.join(os.getcwd(), "temp_processing")
            os.makedirs(temp_dir, exist_ok=True)
            
            local_path = os.path.join(temp_dir, "kolon.xlsx")
            blob.download_to_filename(local_path)
            print(f"kolon.xlsx 템플릿 다운로드 완료: {local_path}")
            return local_path
        except Exception as e:
            print(f"kolon.xlsx 템플릿 다운로드 실패: {e}")
            return None
    
    def get_dept_mapping(self):
        """Firebase에서 부서 매핑 데이터 가져오기"""
        try:
            if self.db is None:
                print("Firebase Firestore 연결이 초기화되지 않았습니다.")
                return {}
            
            # Firestore에서 부서 정보 조회
            docs = self.db.collection('dept_codes').stream()
            
            dept_mapping = {}
            for doc in docs:
                data = doc.to_dict()
                if 'DEPT_NM' in data and 'DEPT_CD' in data:
                    dept_mapping[data['DEPT_NM']] = data['DEPT_CD']
            
            print(f"Firebase에서 {len(dept_mapping)}개의 부서 매핑 정보를 가져왔습니다.")
            return dept_mapping
            
        except Exception as e:
            print(f"Firebase에서 부서 매핑 데이터 조회 실패: {str(e)}")
            return {}
        
    def convert_xls_to_csv(self, xls_file_path):
        """XLS/XLSX/CSV 파일을 CSV로 변환 또는 확인"""
        try:
            print(f"🔍 변환 시작: {os.path.basename(xls_file_path)}")
            print(f"   원본 파일: {xls_file_path}")
            
            # 파일 확장자 확인
            if xls_file_path.endswith('.csv'):
                # 이미 CSV 파일인 경우 변환 필요 없음
                print(f"   이미 CSV 파일입니다")
                try:
                    # CSV 파일 읽어서 데이터 확인만
                    df = pd.read_csv(xls_file_path)
                    print(f"   데이터 형태: {df.shape[0]}행 x {df.shape[1]}열")
                    print(f"   컬럼: {list(df.columns)}")
                    print(f"CSV 파일 확인 완료")
                    return xls_file_path
                except Exception as e:
                    print(f"CSV 파일 읽기 실패: {e}")
                    return None
            elif xls_file_path.endswith('.xlsx'):
                # Excel 파일 처리
                print(f"   Excel 파일로 읽기 시도...")
                df = pd.read_excel(xls_file_path)
                print(f"   데이터 형태: {df.shape[0]}행 x {df.shape[1]}열")
                print(f"   컬럼: {list(df.columns)}")
                csv_path = xls_file_path.replace('.xlsx', '.csv')
                df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                print(f"Excel → .csv 변환 완료: {csv_path}")
                return csv_path
            else:
                # XLS 파일 처리 (HTML 내용인지 확인)
                try:
                    with open(xls_file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    # HTML 내용인지 확인
                    if '<html' in content.lower() or '<table' in content.lower():
                        # HTML 테이블 파싱
                        print(f"   HTML 테이블 형태 감지")
                        df = pd.read_html(xls_file_path)[0]
                        print(f"   데이터 형태: {df.shape[0]}행 x {df.shape[1]}열")
                        print(f"   컬럼: {list(df.columns)}")
                        
                        # CSV로 저장
                        # 원본 파일명에서 확장자만 변경하여 CSV 저장
                        csv_path = os.path.splitext(xls_file_path)[0] + '.csv'
                        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                        print(f"HTML → .csv 변환 완료: {csv_path}")
                        return csv_path
                    else:
                        # 일반 Excel 파일 처리
                        print(f"   일반 Excel 파일로 읽기 시도...")
                        df = pd.read_excel(xls_file_path)
                        print(f"   데이터 형태: {df.shape[0]}행 x {df.shape[1]}열")
                        print(f"   컬럼: {list(df.columns)}")
                        # 원본 파일명에서 확장자만 변경하여 CSV 저장
                        csv_path = os.path.splitext(xls_file_path)[0] + '.csv'
                        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                        print(f"Excel → .csv 변환 완료: {csv_path}")
                        return csv_path
                except UnicodeDecodeError:
                    # 인코딩 오류 시 Excel로 읽기 시도
                    print(f"   인코딩 오류, Excel로 읽기 시도...")
                    df = pd.read_excel(xls_file_path)
                    print(f"   데이터 형태: {df.shape[0]}행 x {df.shape[1]}열")
                    print(f"   컬럼: {list(df.columns)}")
                    csv_path = xls_file_path.replace('.xls', '.csv')
                    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                    print(f"Excel → .csv 변환 완료: {csv_path}")
                    return csv_path
                
        except Exception as e:
            print(f"변환 실패: {xls_file_path}, 오류: {e}")
            return None

    def parse_korean_datetime(self, date_str):
        """한글 날짜를 datetime으로 변환"""
        try:
            # "2025년 7월 30일 오후 10:13" 형식 파싱
            pattern = r'(\d{4})년\s+(\d{1,2})월\s+(\d{1,2})일\s+(오전|오후)\s+(\d{1,2}):(\d{2})'
            match = re.match(pattern, date_str)
            if match:
                year, month, day, ampm, hour, minute = match.groups()
                hour = int(hour)
                if ampm == '오후' and hour < 12:
                    hour += 12
                elif ampm == '오전' and hour == 12:
                    hour = 0
                return datetime(int(year), int(month), int(day), hour, int(minute))
        except Exception as e:
            print(f"날짜 파싱 실패: {date_str}, 오류: {e}")
        return None

    def preprocess_jaegyeong_data(self, df):
        """재경팀 데이터 전처리"""
        try:
            # 날짜 컬럼 전처리
            df["매출일자"] = df["매출일자"].astype(str).str.zfill(8)
            df["승인시각"] = df["승인시각"].astype(str).str.zfill(6)
            df["날짜"] = df["매출일자"] + df["승인시각"]
            df["날짜변환"] = df["날짜"].str.slice(0, 12)
            df["일자"] = df["날짜"].str.slice(0, 8)
            
            # 거래ID 추가
            df["거래ID"] = range(1, len(df) + 1)
            
            # 숫자 데이터 전처리
            numeric_cols = ["매출금액", "현지거래금액", "원화환산금액", "해외접수달러금액"]
            for col in numeric_cols:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce')
            
            print("재경팀 데이터 전처리 완료")
            return df
        except Exception as e:
            print(f"재경팀 데이터 전처리 실패: {e}")
            return None

    def preprocess_openai_data(self, df):
        """OpenAI 데이터 전처리"""
        try:
            # 날짜 전처리
            df["datetime"] = df["날짜"].apply(self.parse_korean_datetime)
            df["날짜변환"] = df["datetime"].dt.strftime("%Y%m%d%H%M")
            df["일자"] = df["datetime"].dt.strftime("%Y%m%d")
            
            # 금액 전처리 (공백 제거 후 $ 제거)
            df["금액(USD)"] = pd.to_numeric(
                df["금액"].str.strip().str.replace('$', '').str.replace(',', ''),
                errors='coerce'
            )
            
            print("✅ OpenAI 데이터 전처리 완료")
            return df
        except Exception as e:
            print(f"❌ OpenAI 데이터 전처리 실패: {e}")
            return None

    def match_data(self, df_jaegyeong, df_openai):
        """두 데이터프레임을 매칭"""
        try:
            # 데이터 병합 (날짜와 금액으로 매칭)
            print("🔄 데이터 매칭 시작...")
            print(f"   재경팀 데이터: {len(df_jaegyeong)}건")
            print(f"   OpenAI 데이터: {len(df_openai)}건")
            
            # 디버깅을 위한 데이터 샘플 출력
            print("📊 재경팀 데이터 샘플:")
            print(f"   일자 샘플: {df_jaegyeong['일자'].head().tolist()}")
            print(f"   해외접수달러금액 샘플: {df_jaegyeong['해외접수달러금액'].head().tolist()}")
            
            print("📊 OpenAI 데이터 샘플:")
            print(f"   일자 샘플: {df_openai['일자'].head().tolist()}")
            print(f"   금액(USD) 샘플: {df_openai['금액(USD)'].head().tolist()}")
            print(f"   계정 샘플: {df_openai['계정'].head().tolist()}")
            
            # 먼저 날짜로만 매칭하고, 그 다음 금액 차이를 확인
            merged = pd.merge(
                df_jaegyeong,
                df_openai[["일자", "금액(USD)", "계정"]],
                how="left",
                left_on="일자",
                right_on="일자"
            )
            
            # 금액 차이 계산 (0.5 USD 이내만 매칭으로 인정)
            merged["금액차이"] = abs(merged["해외접수달러금액"] - merged["금액(USD)"])
            print(f"금액 차이 분석:")
            if len(merged) > 0:
                print(f"   최소 차이: {merged['금액차이'].min():.2f}")
                print(f"   최대 차이: {merged['금액차이'].max():.2f}")
                print(f"   평균 차이: {merged['금액차이'].mean():.2f}")
                print(f"   0.5 USD 이내 건수: {(merged['금액차이'] <= 0.5).sum()}")
            
            # 금액 차이가 0.5 USD 이내이고 계정이 있는 경우만 매칭으로 인정
            matched_condition = (~merged["계정"].isna()) & (merged["금액차이"] <= 0.5)
            matched_data = merged[matched_condition].copy()
            unmatched_data = merged[~matched_condition].copy()
            
            # 중복 제거
            matched_data = matched_data.drop_duplicates(subset=["거래ID"])
            
            print(f"데이터 매칭 완료 (매칭: {len(matched_data)}건, 미매칭: {len(unmatched_data)}건)")
            return matched_data, unmatched_data
        except Exception as e:
            print(f"데이터 매칭 실패: {e}")
            return None, None

    def create_summary_data(self, kolon_df):
        """요약 데이터 생성"""
        try:
            # 주요 월 찾기
            main_month = kolon_df['매출일자'].iloc[0][4:6]
            
            # 날짜 목록 생성
            dates = []
            
            # 이전 월의 데이터 확인
            prev_month_data = kolon_df[kolon_df['매출일자'].str[4:6] != main_month]
            if not prev_month_data.empty:
                # 이전 월의 데이터 추가
                for _, row in prev_month_data.iterrows():
                    dates.append(row['매출일자'])
            
            # 주요 월의 1일부터 말일까지 생성
            year = kolon_df['매출일자'].iloc[0][:4]
            last_day = 31 if main_month in ['01', '03', '05', '07', '08', '10', '12'] else \
                      30 if main_month in ['04', '06', '09', '11'] else \
                      29 if main_month == '02' and int(year) % 4 == 0 else 28
            
            # 주요 월의 모든 날짜 추가
            for day in range(1, last_day + 1):
                dates.append(f"{year}{main_month}{str(day).zfill(2)}")
            
            # 데이터 딕셔너리 생성
            summary_data = {
                '날짜': [],
                '승인금액(USD)': [],
                '승인금액(원화)': []
            }
            
            # 합계 계산
            usd_sum = kolon_df['해외접수달러금액'].sum()
            krw_sum = kolon_df['원화환산금액'].sum()
            
            # 합계 행 추가
            summary_data['날짜'].append('합계')
            summary_data['승인금액(USD)'].append(usd_sum)
            summary_data['승인금액(원화)'].append(krw_sum)
            
            # 날짜별 데이터 행 추가
            for date in dates:
                date_data = kolon_df[kolon_df['매출일자'] == date]
                summary_data['날짜'].append(f"{date[:4]}.{date[4:6]}.{date[6:]}")
                
                if not date_data.empty:
                    summary_data['승인금액(USD)'].append(date_data['해외접수달러금액'].sum())
                    summary_data['승인금액(원화)'].append(date_data['원화환산금액'].sum())
                else:
                    summary_data['승인금액(USD)'].append(0)
                    summary_data['승인금액(원화)'].append(0)
            
            print("요약 데이터 생성 완료")
            return summary_data
            
        except Exception as e:
            print(f"요약 데이터 생성 실패: {e}")
            return None
    
    def filter_kolon_columns(self, kolon_df):
        """코오롱 해외결제 시트용 6개 컬럼 필터링"""
        try:
            # 필요한 6개 컬럼만 선택
            required_columns = [
                '매출일자', 
                '승인시각', 
                '해외접수달러금액', 
                '기준환율', 
                '해외사용수수료', 
                '원화환산금액'
            ]
            
            # 존재하는 컬럼만 선택 (누락된 컬럼이 있을 경우 대비)
            available_columns = [col for col in required_columns if col in kolon_df.columns]
            
            if len(available_columns) != len(required_columns):
                missing_columns = set(required_columns) - set(available_columns)
                print(f"⚠️ 누락된 컬럼: {missing_columns}")
                print(f"📋 사용 가능한 컬럼: {available_columns}")
            
            filtered_df = kolon_df[available_columns].copy()
            print(f"코오롱 해외결제 시트 컬럼 필터링 완료: {len(available_columns)}개 컬럼")
            
            return filtered_df
            
        except Exception as e:
            print(f"컬럼 필터링 실패: {e}")
            print(f"원본 데이터 컬럼: {list(kolon_df.columns)}")
            # 실패 시 원본 데이터 반환
            return kolon_df
    
    def save_kolon_excel(self, kolon_df, output_path):
        """코오롱 전용 Excel 파일 저장"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. 요약 시트 생성
                summary_data = self.create_summary_data(kolon_df)
                if summary_data:
                    # 데이터프레임 생성
                    summary_df = pd.DataFrame(summary_data)
                    
                    # 워크시트 생성
                    worksheet = writer.book.create_sheet("요약", 0)
                    
                    # 헤더 추가 (2행에 추가)
                    worksheet.cell(row=2, column=2, value="")  # B2
                    worksheet.cell(row=2, column=3, value="승인금액(USD)")  # C2
                    worksheet.cell(row=2, column=4, value="승인금액(원화)")  # D2
                    
                    # 데이터 입력 (3행부터)
                    current_row = 3
                    for i in range(len(summary_data['날짜'])):
                        worksheet.cell(row=current_row, column=2, value=summary_data['날짜'][i])
                        worksheet.cell(row=current_row, column=3, value=summary_data['승인금액(USD)'][i])
                        worksheet.cell(row=current_row, column=4, value=summary_data['승인금액(원화)'][i])
                        current_row += 1
                    
                    # 테두리 스타일 정의
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # 데이터가 있는 마지막 행 찾기
                    max_row = len(summary_data['날짜']) + 2  # 헤더 행 + 데이터 행
                    
                    # B,C,D열에 테두리와 스타일 적용
                    for row in range(2, max_row + 1):  # 2행부터 마지막 행까지
                        for col in range(2, 5):  # B(2), C(3), D(4)열
                            cell = worksheet.cell(row=row, column=col)
                            cell.border = thin_border
                            
                            # sum 행에만 배경색 적용
                            if row == 3:  # 3행에 합계가 위치
                                green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                                cell.fill = green_fill
                
                # 2. 코오롱 해외결제 시트 (특정 6개 컬럼만 출력)
                kolon_filtered_df = self.filter_kolon_columns(kolon_df)
                kolon_filtered_df.to_excel(writer, sheet_name="코오롱 해외결제", index=False)
                
            print(f"Excel 파일 저장 완료: {output_path}")
            return True
            
        except Exception as e:
            print(f"Excel 파일 저장 실패: {e}")
            return False
    
    def extract_amount_from_summary(self, excel_path):
        """코오롱_청구내역서 Excel 파일의 요약 시트 D3 셀에서 금액 추출"""
        try:
            workbook = load_workbook(excel_path)
            if '요약' in workbook.sheetnames:
                summary_sheet = workbook['요약']
                amount_cell = summary_sheet.cell(row=3, column=4)  # D3 셀
                amount = amount_cell.value
                
                if amount and isinstance(amount, (int, float)):
                    print(f"요약 시트 D3 셀에서 금액 추출: {amount:,}원")
                    return float(amount)
                else:
                    print(f"D3 셀에서 유효한 금액을 찾을 수 없습니다: {amount}")
                    return None
            else:
                print("요약 시트를 찾을 수 없습니다")
                return None
        except Exception as e:
            print(f"금액 추출 실패: {e}")
            return None
        finally:
            if 'workbook' in locals():
                workbook.close()
    
    def calculate_amount_without_vat(self, total_amount):
        """부가세 10%를 제외한 금액 계산 (1원 오차 보정 포함)"""
        try:
            # 부가세 포함 금액에서 부가세 제외
            # 총금액 = 공급가액 + 부가세(공급가액의 10%)
            # 총금액 = 공급가액 × 1.1
            # 공급가액 = 총금액 / 1.1
            amount_without_vat = round(total_amount / 1.1)
            
            # 검증: 부가세 제외 금액 + 부가세 = 실제 고지서 금액인지 확인
            calculated_total = round(amount_without_vat * 1.1)
            difference = total_amount - calculated_total
            
            if difference != 0:
                # 1원 차이가 있으면 부가세 제외 금액을 보정
                amount_without_vat += difference
                print(f"1원 오차 보정: {difference:+d}원 조정")
                
            # 최종 검증
            final_total = round(amount_without_vat * 1.1)
            print(f"부가세 제외 계산: {total_amount:,}원 → {amount_without_vat:,}원 (검증: {final_total:,}원)")
            
            return amount_without_vat
        except Exception as e:
            print(f"부가세 제외 계산 실패: {e}")
            return None
    
    def calculate_meta_ics_usage(self, collection_date):
        """Meta ICS 사용량 자동 계산 (코오롱용)"""
        try:
            date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            days_in_month = calendar.monthrange(date_obj.year, date_obj.month)[1]
            
            # 코오롱 계정 정보 (고정값)
            metalcs_accounts = 65  # D39 셀의 계정 수 (코오롱은 65개 계정 사용)
            ssl_vpn_accounts = 0   # SSL-VPN은 사용하지 않음
            
            # 사용일수 계산 (매일 사용 가정)
            metalcs_usage_days = metalcs_accounts * days_in_month
            ssl_vpn_usage_days = ssl_vpn_accounts * days_in_month
            
            # 라이선스 계산
            metalcs_licenses = metalcs_usage_days / days_in_month
            ssl_vpn_licenses = ssl_vpn_usage_days / days_in_month
            
            print(f"코오롱 Meta ICS 사용량 계산:")
            print(f"   - 해당 월 일수: {days_in_month}일")
            print(f"   - MetaLCS 사용일수: {metalcs_usage_days}일 ({metalcs_accounts}계정 × {days_in_month}일)")
            print(f"   - SSL-VPN 사용일수: {ssl_vpn_usage_days}일 ({ssl_vpn_accounts}계정 × {days_in_month}일)")
            print(f"   - MetaLCS 라이선스: {metalcs_licenses}개")
            print(f"   - SSL-VPN 라이선스: {ssl_vpn_licenses}개")
            
            return {
                'days_in_month': days_in_month,
                'metalcs_usage_days': metalcs_usage_days,
                'ssl_vpn_usage_days': ssl_vpn_usage_days,
                'metalcs_licenses': metalcs_licenses,
                'ssl_vpn_licenses': ssl_vpn_licenses
            }
        except Exception as e:
            print(f"Meta ICS 사용량 계산 실패: {e}")
            return None
    
    def update_kolon_template(self, template_path, amount_without_vat, collection_date, total_amount):
        """kolon.xlsx 템플릿 파일 업데이트"""
        try:
            date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            year_month = f"{date_obj.year}년 {date_obj.month:02d}월"
            date_prefix = f"{str(date_obj.year)[2:]}{date_obj.month:02d}"
            
            # 출력 파일명 생성
            output_filename = f"{date_prefix}_코오롱FnC_상담솔루션 청구내역서.xlsx"
            output_path = os.path.join(self.download_dir, output_filename)
            
            # 템플릿 파일 복사
            shutil.copy2(template_path, output_path)
            print(f"템플릿 파일 복사 완료: {output_filename}")
            
            workbook = load_workbook(output_path)
            
            # B9 셀에 문서번호 설정 (MMP-{년월} 형식)
            document_number = f"MMP-{date_prefix}"
            for sheet in workbook.worksheets:
                if '세부내역' in sheet.title or '대외공문' in sheet.title:
                    sheet.cell(row=9, column=2).value = f"문서번호  : {document_number}"
                    print(f"{sheet.title} B9 셀에 문서번호 설정 완료: {document_number}")
            
            # 1. 대외공문 시트 업데이트
            if '대외공문' in workbook.sheetnames:
                doc_sheet = workbook['대외공문']
                
                # B13 셀 업데이트: "제       목 : 2025년 07월 상담솔루션 서비스 수수료 정산 요청"
                b13_cell = doc_sheet.cell(row=13, column=2)
                if b13_cell.value and isinstance(b13_cell.value, str):
                    old_text = b13_cell.value
                    # 년월 패턴을 찾아서 교체
                    new_text = re.sub(r'\d{4}년 \d{1,2}월', year_month, old_text)
                    b13_cell.value = new_text
                    print(f"대외공문 B13 셀 업데이트: {old_text} → {new_text}")
                
                # B16 셀 업데이트 (B,C,D,E,F,G 16행 병합)
                b16_cell = doc_sheet.cell(row=16, column=2)
                if b16_cell.value and isinstance(b16_cell.value, str):
                    old_text = b16_cell.value
                    new_text = re.sub(r'2025년\d{1,2}월', year_month, old_text)
                    if new_text == old_text:
                        new_text = re.sub(r'\d{4}년\d{1,2}월', year_month, old_text)
                    b16_cell.value = new_text
                    print(f"대외공문 B16 셀 업데이트: {old_text} → {new_text}")
                
                # 하단 테이블 수식 업데이트 (B24, D24, D25)
                self.update_formula_references(doc_sheet, year_month)
            
            # 2. 시트명 변경 및 해당 시트의 B,C,D,E1 병합 셀 텍스트 업데이트
            for sheet in workbook.worksheets:
                if re.match(r'\d{4}년 \d{1,2}월', sheet.title):
                    old_title = sheet.title
                    sheet.title = year_month
                    print(f"시트명 변경: {old_title} → {year_month}")
                    
                    # B,C,D,E1 병합 셀의 텍스트 업데이트 ("2025년 07월 수수료 청구 금액" 형태)
                    b1_cell = sheet.cell(row=1, column=2)  # B1 셀
                    if b1_cell.value and isinstance(b1_cell.value, str):
                        old_text = b1_cell.value
                        # 년월 패턴을 찾아서 교체
                        new_text = re.sub(r'\d{4}년 \d{1,2}월', year_month, old_text)
                        if new_text != old_text:
                            b1_cell.value = new_text
                            print(f"{year_month} 시트 B1 셀 텍스트 업데이트: {old_text} → {new_text}")
                    break
            
            # 로고 이미지 삽입 (B2 셀)
            try:
                logo_path = os.path.join(os.path.dirname(__file__), 'assets', 'logo.png')
                if os.path.exists(logo_path):
                    img = Image(logo_path)
                    if '대외공문' in workbook.sheetnames:
                        doc_sheet = workbook['대외공문']
                        doc_sheet.add_image(img, 'B2')
                        print("대외공문 시트 B2 셀에 로고 이미지 삽입 완료")
                    elif workbook.worksheets:
                        workbook.worksheets[0].add_image(img, 'B2')
                        print(f"{workbook.worksheets[0].title} 시트 B2 셀에 로고 이미지 삽입 완료")
            except Exception as e:
                print(f"로고 이미지 삽입 실패: {e}")
            
            # 3. 세부내역 시트 업데이트
            if '세부내역' in workbook.sheetnames:
                detail_sheet = workbook['세부내역']
                # E12 셀에 부가세 제외 금액 입력
                detail_sheet.cell(row=12, column=5).value = amount_without_vat
                print(f"세부내역 시트 E12 셀 업데이트: {amount_without_vat:,}원")
                
                # E15 셀에 총금액 비용 그대로 입력 (1.1 계산 전 총금액)
                detail_sheet.cell(row=15, column=5).value = total_amount
                print(f"세부내역 시트 E15 셀 업데이트: {total_amount:,}원 (총금액)")
                
                # Meta ICS 계산 및 업데이트
                ics_data = self.calculate_meta_ics_usage(collection_date)
                if ics_data:
                    # E36 셀에 해당 월 일수 (E38 셀의 =E36*D38 수식이 자동으로 계산됨)
                    detail_sheet.cell(row=36, column=5).value = ics_data['days_in_month']
                    print(f"세부내역 시트 E36 셀 업데이트: {ics_data['days_in_month']}일")
                    
                    print(f"Meta ICS 계산 완료: {ics_data['days_in_month']}일 기준")
            
            # 파일 저장
            workbook.save(output_path)
            workbook.close()
            
            print(f"kolon.xlsx 템플릿 업데이트 완료: {output_filename}")
            return output_path
            
        except Exception as e:
            print(f"kolon.xlsx 템플릿 업데이트 실패: {e}")
            return None
    
    def update_formula_references(self, doc_sheet, year_month):
        """대외공문 시트의 수식에서 시트명 참조 업데이트"""
        try:
            # 수식이 있을 수 있는 셀들 확인 (24행, 25행 등)
            cells_to_check = [
                (24, 2), (24, 3), (24, 4),  # B24, C24, D24
                (25, 2), (25, 3), (25, 4),  # B25, C25, D25
            ]
            
            for row, col in cells_to_check:
                cell = doc_sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    old_formula = cell.value
                    # 수식에서 시트명 참조 패턴 찾아서 교체
                    new_formula = re.sub(r"'(\d{4}년 \d{1,2}월)'!", f"'{year_month}'!", old_formula)
                    if new_formula != old_formula:
                        cell.value = new_formula
                        print(f"대외공문 {chr(64+col)}{row} 셀 수식 업데이트: {old_formula} → {new_formula}")
                        
        except Exception as e:
            print(f"수식 참조 업데이트 오류: {e}")

    def process_kolon_data(self, collection_date, input_dir="temp_processing", selected_filenames=None):
        """코오롱 데이터 전처리 메인 함수"""
        try:
            print("코오롱 데이터 전처리 시작")
            
            temp_dir = input_dir
            if not os.path.exists(temp_dir):
                print(f"입력 폴더를 찾을 수 없습니다: {temp_dir}")
                return False
            
            kolon_files = []
            if selected_filenames:
                # 프론트에 표시된 업로드 파일 목록을 우선 사용
                for filename in selected_filenames:
                    if not filename:
                        continue
                    file_path = os.path.join(temp_dir, filename)
                    if os.path.exists(file_path) and filename.endswith((".xls", ".xlsx", ".csv")):
                        kolon_files.append(file_path)
            else:
                for filename in os.listdir(temp_dir):
                    if "코오롱" in filename and filename.endswith((".xls", ".xlsx", ".csv")):
                        file_path = os.path.join(temp_dir, filename)
                        kolon_files.append(file_path)
            
            if len(kolon_files) != 2:
                print(f"코오롱 파일이 2개 필요합니다. 현재: {len(kolon_files)}개")
                return False
            
            # 2. CSV 변환
            csv_files = []
            for file_path in kolon_files:
                csv_path = self.convert_xls_to_csv(file_path)
                if csv_path:
                    csv_files.append(csv_path)
            
            if len(csv_files) != 2:
                print("CSV 변환 실패")
                return False
            
            print(f"변환 완료: {len(csv_files)}개 파일")
            
            # 3. 데이터 읽기
            print("변환된 CSV 파일 목록:")
            for f in csv_files:
                print(f"   - {f}")
                
            # 파일 찾기 (파일 내용으로 구분)
            jaegyeong_files = []
            openai_files = []
            
            for f in csv_files:
                try:
                    df = pd.read_csv(f)
                    if '매출일자' in df.columns:  # 재경팀 데이터
                        jaegyeong_files.append(f)
                    elif '날짜' in df.columns:  # OpenAI 데이터
                        openai_files.append(f)
                except Exception as e:
                    print(f"파일 읽기 실패: {f}, 오류: {e}")
            
            if not jaegyeong_files:
                print("재경팀 데이터 파일을 찾을 수 없습니다")
                return False
            if not openai_files:
                print("OpenAI 데이터 파일을 찾을 수 없습니다")
                return False
                
            jaegyeong_file = jaegyeong_files[0]
            openai_file = openai_files[0]
            
            # 재경팀 데이터 로드 (빈 행 제거)
            df_jaegyeong = pd.read_csv(jaegyeong_file).dropna(how='all')
            
            # OpenAI 데이터 로드
            df_openai = pd.read_csv(openai_file)
            
            print(f"데이터 로드 완료 (재경팀: {len(df_jaegyeong)}행, OpenAI: {len(df_openai)}행)")
            
            # 4. 데이터 전처리
            df_jaegyeong = self.preprocess_jaegyeong_data(df_jaegyeong)
            if df_jaegyeong is None:
                return False
                
            df_openai = self.preprocess_openai_data(df_openai)
            if df_openai is None:
                return False
            
            print("데이터 전처리 완료")
            
            # 5. 데이터 매칭 (모든 OpenAI 데이터와 매칭)
            matched_data, unmatched_data = self.match_data(df_jaegyeong, df_openai)
            if matched_data is None:
                return False
            
            print(f"데이터 매칭 완료 (매칭: {len(matched_data)}건, 미매칭: {len(unmatched_data)}건)")
            
            # 7. 결과 파일 생성
            # 7.1. 매칭 결과 CSV 저장
            date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            collection_month = datetime.strptime(collection_date, '%Y-%m-%d').strftime('%Y%m')
            
            # 1. OpenAI 매칭결과 파일 생성 (모든 계정의 매칭된 데이터)
            openai_filename = f"OpenAI_정확매칭결과_{date_str}_{collection_month}.csv"
            openai_path = os.path.join(self.download_dir, openai_filename)
            
            # 모든 매칭된 데이터로부터 결과 생성
            if len(matched_data) > 0:
                # Firebase에서 부서 매핑 정보 가져오기
                dept_mapping = self.get_dept_mapping()
                
                # 프로젝트 컬럼 설정 (부서 매핑이 있으면 사용, 없으면 계정값 사용)
                if dept_mapping:
                    project_values = matched_data['계정'].map(dept_mapping).fillna(matched_data['계정'])
                    matched_count = project_values.ne(matched_data['계정']).sum()
                    print(f"부서 매핑 결과: {matched_count}/{len(matched_data)} 건 매칭됨")
                else:
                    project_values = matched_data['계정']
                    print("Firebase 연결 실패로 프로젝트 컬럼을 계정값으로 설정합니다.")
                
                # 기존 시스템과 동일한 컬럼 순서로 DataFrame 생성 (모든 계정)
                result_df = pd.DataFrame({
                    '매출일자': matched_data['매출일자'].astype(str),  # YYYYMMDD 형식 유지
                    '승인시각': matched_data['승인시각'].astype(str).str.zfill(6),  # HHMMSS 형식
                    '매출금액': matched_data['원화환산금액'],  # 원화 금액 사용
                    '계정': matched_data['계정'],  # 매칭된 계정값
                    '표준적요': 156,  # 고정값
                    '증빙유형': '003',  # 고정값
                    '적요': 'OpenAI_GPT API 토큰 비용_' + matched_data['계정'],  # 계정별 적요
                    '프로젝트': project_values  # 부서 매핑된 값 또는 계정값
                })
                
                # 컬럼 순서 정리 (기존 시스템과 동일)
                column_order = ['매출일자', '승인시각', '매출금액', '계정', '표준적요', '증빙유형', '적요', '프로젝트']
                result_df = result_df[column_order]
                
                result_df.to_csv(openai_path, index=False, encoding='utf-8-sig')
                print(f"OpenAI 매칭결과 CSV 저장 완료: {openai_filename} (총 {len(result_df)}건, 모든 계정 포함)")
                
                # 코오롱 계정만 필터링 (코오롱 전용 파일 생성을 위해)
                kolon_matched_data = matched_data[matched_data['계정'] == '코오롱'].copy()
                print(f"코오롱 계정 매칭 결과: {len(kolon_matched_data)}건")
            else:
                print("매칭된 데이터가 없어 OpenAI 매칭결과 파일을 생성할 수 없습니다")
                return False
            
            # 2. 코오롱 전용 요약 보고서 Excel 저장
            if len(kolon_matched_data) > 0:
                report_filename = f"코오롱_청구내역서_{date_str}_{collection_month}.xlsx"
                report_path = os.path.join(self.download_dir, report_filename)
                
                if self.save_kolon_excel(kolon_matched_data, report_path):
                    print(f"청구내역서 Excel 저장 완료: {report_filename}")
                    
                    # 3. Firebase에서 kolon.xlsx 템플릿 다운로드 및 청구서 생성
                    print("Firebase kolon.xlsx 템플릿 처리 시작")
                
                    # 3.1. 요약 시트에서 금액 추출
                    total_amount = self.extract_amount_from_summary(report_path)
                    if total_amount is None:
                        print("요약 시트에서 금액 추출 실패")
                        return False
                    
                    # 3.2. 부가세 제외 금액 계산
                    amount_without_vat = self.calculate_amount_without_vat(total_amount)
                    if amount_without_vat is None:
                        print("부가세 제외 계산 실패")
                        return False
                    
                    # 3.3. kolon.xlsx 템플릿 다운로드
                    template_path = self.download_kolon_template()
                    if template_path is None:
                        print("kolon.xlsx 템플릿 다운로드 실패")
                        return False
                    
                    # 3.4. 템플릿 업데이트 및 청구서 생성
                    final_invoice_path = self.update_kolon_template(template_path, amount_without_vat, collection_date, total_amount)
                    if final_invoice_path is None:
                        print("코오롱 청구서 생성 실패")
                        return False
                    
                    print(f"코오롱 전처리 완료! 총 3개 파일 생성:")
                    print(f"   1. OpenAI 매칭결과: {openai_filename} (모든 계정)")
                    print(f"   2. 코오롱 청구내역서: {report_filename}")
                    print(f"   3. 코오롱FnC 상담솔루션 청구내역서: {os.path.basename(final_invoice_path)}")
                    
                    # self.cleanup_temp_folder()
                    
                    return True
                else:
                    print("청구내역서 Excel 저장 실패")
                    return False
            else:
                print("코오롱 계정 매칭 데이터가 없어 코오롱 전용 파일들을 생성하지 않습니다.")
                print(f"OpenAI 매칭결과만 생성 완료: {openai_filename} (모든 계정)")
                
                # self.cleanup_temp_folder()
                return True
            
        except Exception as e:
            import traceback
            print(f"전처리 실패: {e}")
            print(f"상세 에러: {traceback.format_exc()}")
            return False
    
    def cleanup_temp_folder(self):
        """temp_processing 폴더 정리"""
        try:
            temp_dir = "temp_processing"
            if os.path.exists(temp_dir):
                # 모든 파일 삭제
                for filename in os.listdir(temp_dir):
                    file_path = os.path.join(temp_dir, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                        print(f"🗑️ 삭제: {filename}")
                
                print("temp_processing 폴더 정리 완료")
            else:
                print("temp_processing 폴더가 존재하지 않습니다")
        except Exception as e:
            print(f"폴더 정리 실패: {e}")

