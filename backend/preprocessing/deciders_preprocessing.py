import os
import re
import shutil
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import firebase_admin
from firebase_admin import credentials, storage
from backend.utils.secrets_manager import get_firebase_secret

class DecidersPreprocessor:
    def __init__(self):
        self.download_dir = os.path.join(os.getcwd(), "temp_processing")
        os.makedirs(self.download_dir, exist_ok=True)
        self.setup_firebase()
        
        # 발신번호별 매핑
        self.sender_mapping = {
            "18005073": "엑스퍼",
            "16610581": "스마트웰컴", 
            "16881635": "바이오숨"
        }
        
        # 청구서 매핑
        self.invoice_mapping = {
            "엑스퍼": "디싸이더스",  # 디싸이더스 청구서
            "스마트웰컴": "애드프로젝트",  # 애드프로젝트 청구서
            "바이오숨": "애드프로젝트"  # 애드프로젝트 청구서
        }
        
    def setup_firebase(self):
        """Firebase Storage 연결 설정"""
        try:
            cred_dict = get_firebase_secret()
            BUCKET_NAME = "services-e42af.firebasestorage.app"
            
            # 기존 앱이 있으면 삭제하고 새로 초기화
            if firebase_admin._apps:
                firebase_admin.delete_app(firebase_admin.get_app())
            
            firebase_admin.initialize_app(
                credentials.Certificate(cred_dict),
                {"storageBucket": BUCKET_NAME}
            )
            
            self.bucket = storage.bucket()
            print("Firebase Storage 연결 완료")
        except Exception as e:
            print(f"Firebase 연결 실패: {e}")
            self.bucket = None
    
    def convert_all_to_csv(self):
        """temp_processing 폴더의 모든 디싸이더스애드프로젝트 파일을 CSV로 변환"""
        temp_dir = "temp_processing"
        csv_files = []
        
        if os.path.exists(temp_dir):
            for filename in os.listdir(temp_dir):
                if "디싸이더스애드프로젝트" in filename:
                    file_path = os.path.join(temp_dir, filename)
                    
                    if filename.endswith((".xlsx", ".xls")):
                        try:
                            df = pd.read_excel(file_path)
                            csv_filename = filename.replace('.xlsx', '.csv').replace('.xls', '.csv')
                            csv_path = os.path.join(temp_dir, csv_filename)
                            
                            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                            csv_files.append(csv_path)
                            print(f"CSV 변환 완료: {csv_filename}")
                            
                        except Exception as e:
                            print(f"CSV 변환 실패: {filename}, 오류: {e}")
                    
                    elif filename.endswith(".csv"):
                        csv_files.append(file_path)
                        print(f"CSV 파일 확인: {filename}")
        
        return csv_files
    
    def merge_sms_files(self, csv_files):
        """발송이력 키워드가 포함된 파일들을 병합"""
        import unicodedata
        sms_files = []
        other_files = []
        
        for csv_file in csv_files:
            filename = os.path.basename(csv_file)
            normalized_filename = unicodedata.normalize('NFC', filename)
            
            # 발송이력 관련 파일이면서 채팅진행건리스트가 아닌 경우
            contains_sms = any(word in normalized_filename for word in ["발송이력", "발송", "이력"])
            is_chat_list = "채팅진행건리스트" in normalized_filename
            
            if contains_sms and not is_chat_list:
                sms_files.append(csv_file)
            else:
                other_files.append(csv_file)
        
        if not sms_files:
            print("발송이력 파일을 찾을 수 없습니다")
            return None, other_files
        
        # 발송이력 파일들 병합
        merged_df = pd.DataFrame()
        
        for sms_file in sms_files:
            try:
                df = pd.read_csv(sms_file, encoding='utf-8-sig')
                merged_df = pd.concat([merged_df, df], ignore_index=True)
                print(f"병합: {os.path.basename(sms_file)} ({len(df)}행)")
            except Exception as e:
                print(f"병합 실패: {sms_file}, 오류: {e}")
        
        if not merged_df.empty:
            # 병합된 파일 저장
            merged_path = os.path.join("temp_processing", "merged_sms_data.csv")
            merged_df.to_csv(merged_path, index=False, encoding='utf-8-sig')
            print(f"발송이력 병합 완료: {len(merged_df)}행")
            return merged_path, other_files
        
        return None, other_files
    
    def analyze_sender_numbers(self, merged_sms_path):
        """발신번호 분석 및 미지의 번호 감지"""
        try:
            df = pd.read_csv(merged_sms_path, encoding='utf-8-sig')
            
            if '발신번호' not in df.columns:
                print("발신번호 컬럼을 찾을 수 없습니다")
                return None
            
            # 발신번호를 문자열로 변환하고 정리
            df['발신번호_정리'] = df['발신번호'].astype(str).str.replace('.0', '')
            
            # 고유 발신번호 확인
            unique_senders = df['발신번호_정리'].unique()
            unknown_senders = []
            
            for sender in unique_senders:
                if sender not in self.sender_mapping and sender != 'nan':
                    unknown_senders.append(sender)
            
            # 미지의 번호가 있으면 알림
            if unknown_senders:
                print(f"다른번호 감지: {', '.join(unknown_senders)} - 스마트웰 or 유리제로")
                # 실제 환경에서는 팝업으로 표시해야 함
            
            return df
            
        except Exception as e:
            print(f"발신번호 분석 실패: {e}")
            return None
    
    def count_message_types_by_sender(self, df):
        """발신번호별로 문자유형 카운트 (성공 전달만)"""
        # 성공(전달)인 경우만 필터링
        success_df = df[df['발송상태'] == '성공(전달)']
        
        results = {
            "디싸이더스": {"SMS": 0, "LMS": 0, "MMS": 0, "TALK": 0},
            "애드프로젝트": {"SMS": 0, "LMS": 0, "MMS": 0, "TALK": 0}
        }
        
        for _, row in success_df.iterrows():
            sender = str(row['발신번호_정리'])
            msg_type = row.get('문자유형', '')
            
            # 발신번호로 브랜드 확인
            brand = self.sender_mapping.get(sender, "기타")
            
            # 청구서 타입 결정
            if brand == "엑스퍼":
                invoice_type = "디싸이더스"
            elif brand in ["스마트웰컴", "바이오숨"]:
                invoice_type = "애드프로젝트"
            else:
                # 미지의 번호는 애드프로젝트로 분류 (기본값)
                invoice_type = "애드프로젝트"
            
            # 문자유형별 카운트
            if msg_type == "SMS":
                results[invoice_type]["SMS"] += 1
            elif msg_type in ["LMS", "LMS/MMS"]:
                results[invoice_type]["LMS"] += 1
            elif msg_type in ["MMS"]:
                results[invoice_type]["MMS"] += 1
            elif msg_type in ["TALK", "TALK(알림톡)"]:
                results[invoice_type]["TALK"] += 1
        
        return results
    
    def download_template(self, template_name):
        """Firebase에서 템플릿 다운로드"""
        if not self.bucket:
            return None
        
        try:
            blob = self.bucket.blob(template_name)
            temp_dir = os.path.join(os.getcwd(), "temp_processing")
            os.makedirs(temp_dir, exist_ok=True)
            
            local_path = os.path.join(temp_dir, template_name)
            blob.download_to_filename(local_path)
            
            return local_path
        except Exception as e:
            print(f"템플릿 다운로드 실패: {e}")
            return None
    
    def process_chat_list_file(self, csv_files):
        """채팅진행건리스트 파일 처리 - 디싸이더스/애드프로젝트 카카오 채팅 카운트"""
        import unicodedata
        try:
            deciders_chat_count = 0
            adproject_chat_count = 0
            
            for csv_file in csv_files:
                normalized_filename = unicodedata.normalize('NFC', os.path.basename(csv_file))
                
                if "채팅진행건리스트" in normalized_filename:
                    print(f"채팅진행건리스트 파일 처리: {os.path.basename(csv_file)}")
                    
                    df = pd.read_csv(csv_file, encoding='utf-8-sig')
                    
                    # 채널명, 채팅유형 컬럼 찾기
                    channel_col = None
                    chat_type_col = None
                    
                    for col in df.columns:
                        norm_col = unicodedata.normalize('NFC', col)
                        if "채널명" in norm_col or "채널" in norm_col:
                            channel_col = col
                        if "채팅유형" in norm_col or "유형" in norm_col:
                            chat_type_col = col
                    
                    if channel_col and chat_type_col:
                        # 디싸이더스: 엑스퍼 + 카카오
                        deciders_kakao = df[(df[channel_col] == '엑스퍼') & (df[chat_type_col] == '카카오')]
                        deciders_count = len(deciders_kakao)
                        deciders_chat_count += deciders_count
                        
                        # 애드프로젝트: 엑스퍼가 아닌 것 + 카카오
                        adproject_kakao = df[(df[channel_col] != '엑스퍼') & (df[chat_type_col] == '카카오')]
                        adproject_count = len(adproject_kakao)
                        adproject_chat_count += adproject_count
                        
                        print(f"   디싸이더스 카카오 채팅: {deciders_count}건")
                        print(f"   애드프로젝트 카카오 채팅: {adproject_count}건")
                    else:
                        print(f"   필요한 컬럼을 찾을 수 없습니다.")
            

            return deciders_chat_count, adproject_chat_count
            
        except Exception as e:
            print(f" 채팅진행건리스트 처리 실패: {e}")
            return 0, 0

    def update_template_with_counts(self, template_path, counts, collection_date, invoice_type, chat_count=0):
        """템플릿에 카운트 데이터 입력 및 파일명 업데이트"""
        try:
            # 날짜 정보 추출
            date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            date_prefix = f"{str(date_obj.year)[2:]}{date_obj.month:02d}"
            
            # 출력 파일명 결정
            if invoice_type == "디싸이더스":
                output_filename = f"{date_prefix}_디싸이더스_상담솔루션 청구내역서.xlsx"
            else:
                output_filename = f"{date_prefix}_애드프로젝트_상담솔루션 청구내역서.xlsx"
            
            download_dir = "temp_processing"
            os.makedirs(download_dir, exist_ok=True)
            output_path = os.path.join(download_dir, output_filename)
            
            # 템플릿 복사
            shutil.copy2(template_path, output_path)
            
            workbook = load_workbook(output_path)
            
            # 수집 날짜로부터 년월 정보 생성
            year_month = f"{date_obj.year}년 {date_obj.month:02d}월"
            
            # 시트명 변경 (예: "2025년 08월" → 수집한 달로 변경)
            for sheet in workbook.worksheets:
                if re.match(r'\d{4}년 \d{1,2}월', sheet.title):
                    old_title = sheet.title
                    sheet.title = year_month
                    print(f"{invoice_type} 시트명 변경: {old_title} → {year_month}")
                    break
            
            # B9 셀에 문서번호 설정 (MMP-{년월} 형식)
            document_number = f"MMP-{date_prefix}"
            for sheet in workbook.worksheets:
                if '세부내역' in sheet.title or '대외공문' in sheet.title:
                    sheet.cell(row=9, column=2).value = f"문서번호  : {document_number}"
                    print(f"{sheet.title} B9 셀에 문서번호 설정 완료: {document_number}")
            
            # 세부내역 시트에서 데이터 입력
            if '세부내역' in workbook.sheetnames:
                sheet = workbook['세부내역']
                
                # D22-D25 셀에 카운트 입력
                sheet['D22'] = counts["SMS"]    # SMS
                sheet['D23'] = counts["LMS"]    # LMS  
                sheet['D24'] = counts["MMS"]    # MMS
                sheet['D25'] = counts["TALK"]   # TALK
                
                # 디싸이더스와 애드프로젝트 모두 D14 셀에 카카오 채팅 카운트 입력
                if chat_count > 0:
                    sheet['D14'] = chat_count
                    print(f"{invoice_type} 카카오 채팅 카운트 입력 - D14: {chat_count}건")
                
                print(f"{invoice_type} 카운트 입력 - SMS:{counts['SMS']}, LMS:{counts['LMS']}, MMS:{counts['MMS']}, TALK:{counts['TALK']}")
            
            # 로고 이미지 삽입 (B2 셀)
            try:
                logo_path = os.path.join(os.path.dirname(__file__), 'assets', 'logo.png')
                if os.path.exists(logo_path):
                    img = Image(logo_path)
                    if '대외공문' in workbook.sheetnames:
                        doc_sheet = workbook['대외공문']
                        doc_sheet.add_image(img, 'B2')
                        print("대외공문 시트 B2 셀에 로고 이미지 삽입 완료")
                    elif workbook.worksheets:
                        workbook.worksheets[0].add_image(img, 'B2')
                        print(f"{workbook.worksheets[0].title} 시트 B2 셀에 로고 이미지 삽입 완료")
            except Exception as e:
                print(f"로고 이미지 삽입 실패: {e}")
            
            # 파일 저장
            workbook.save(output_path)
            workbook.close()
            
            return output_filename
            
        except Exception as e:
            print(f" 템플릿 업데이트 실패: {e}")
            return None
    
    def cleanup_temp_folder(self):
        """temp_processing 폴더 정리"""
        try:
            temp_dir = "temp_processing"
            if os.path.exists(temp_dir):
                for filename in os.listdir(temp_dir):
                    if "디싸이더스애드프로젝트" in filename or "merged_sms_data" in filename:
                        file_path = os.path.join(temp_dir, filename)
                        try:
                            os.remove(file_path)
                        except:
                            pass
                print("temp_processing 폴더 정리 완료")
        except Exception as e:
            print(f"폴더 정리 실패: {e}")
    
    def process_deciders_data(self, collection_date):
        """디싸이더스/애드프로젝트 데이터 전처리 메인 함수"""
        try:
            print("디싸이더스/애드프로젝트 데이터 전처리 시작")
            
            # 1. 모든 파일을 CSV로 변환
            csv_files = self.convert_all_to_csv()
            if not csv_files:
                print("변환할 파일이 없습니다")
                return False
            
            # 2. 발송이력 파일들 병합
            merged_sms_path, other_files = self.merge_sms_files(csv_files)
            if not merged_sms_path:
                print("발송이력 파일 병합 실패")
                return False
            
            # 3. 발신번호 분석
            df = self.analyze_sender_numbers(merged_sms_path)
            if df is None:
                print("발신번호 분석 실패")
                return False
            
            # 4. 문자유형별 카운트
            counts = self.count_message_types_by_sender(df)
            
            # 5. 채팅진행건리스트 파일 처리 (디싸이더스/애드프로젝트 카카오 채팅 카운트)
            deciders_chat_count, adproject_chat_count = self.process_chat_list_file(csv_files)
            
            # 6. 템플릿 다운로드 및 처리
            templates = ["deciders.xlsx", "Adproject.xlsx"]
            processed_files = []
            
            for template_name in templates:
                template_path = self.download_template(template_name)
                if not template_path:
                    continue
                
                # 템플릿별 처리
                if template_name == "deciders.xlsx":
                    invoice_type = "디싸이더스"
                    template_counts = counts["디싸이더스"]
                    template_chat_count = deciders_chat_count  # 디싸이더스 채팅 카운트
                else:  # Adproject.xlsx
                    invoice_type = "애드프로젝트"
                    template_counts = counts["애드프로젝트"]
                    template_chat_count = adproject_chat_count  # 애드프로젝트 채팅 카운트
                
                # 템플릿 업데이트
                output_file = self.update_template_with_counts(
                    template_path, template_counts, collection_date, invoice_type, template_chat_count
                )
                
                if output_file:
                    processed_files.append(output_file)
            
            # self.cleanup_temp_folder()
            
            print(f"전처리 완료: {len(processed_files)}개 파일 생성")
            return processed_files
            
        except Exception as e:
            print(f"전처리 실패: {e}")
            return False
