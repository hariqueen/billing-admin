import os
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import re
import firebase_admin
from firebase_admin import credentials, storage
from backend.utils.secrets_manager import get_firebase_secret
import calendar

class SKPreprocessor:
    def __init__(self):
        self.download_dir = os.path.join(os.getcwd(), "temp_processing")
        os.makedirs(self.download_dir, exist_ok=True)
        self.setup_firebase()
    
    def setup_firebase(self):
        """Firebase Storage 연결 설정"""
        try:
            cred_dict = get_firebase_secret()
            BUCKET_NAME = "services-e42af.firebasestorage.app"
            
            # 기존 앱이 있으면 삭제하고 새로 초기화
            if firebase_admin._apps:
                firebase_admin.delete_app(firebase_admin.get_app())
            
            firebase_admin.initialize_app(
                credentials.Certificate(cred_dict),
                {"storageBucket": BUCKET_NAME}
            )
            
            self.bucket = storage.bucket()
            print("Firebase Storage 연결 완료")
        except Exception as e:
            print(f"Firebase 연결 실패: {e}")
            self.bucket = None
    
    def download_sk_template(self):
        """Firebase에서 skelectlink.xlsx 템플릿 다운로드"""
        if not self.bucket:
            print("Firebase 연결이 없습니다")
            return None
        
        try:
            blob = self.bucket.blob("skelectlink.xlsx")
            temp_dir = os.path.join(os.getcwd(), "temp_processing")
            os.makedirs(temp_dir, exist_ok=True)
            
            local_path = os.path.join(temp_dir, "skelectlink.xlsx")
            blob.download_to_filename(local_path)
            print(f"skelectlink.xlsx 템플릿 다운로드 완료: {local_path}")
            return local_path
        except Exception as e:
            print(f"skelectlink.xlsx 템플릿 다운로드 실패: {e}")
            return None
    
    def get_bill_amount(self, company_name="SK일렉링크"):
        """고지서에서 업데이트된 금액 조회"""
        try:
            import requests
            response = requests.get('http://localhost:5001/api/bill-amounts')
            if response.ok:
                bill_data = response.json()
                if company_name in bill_data:
                    amount_str = bill_data[company_name].get('amount', '')
                    # "123,456원" 형태에서 숫자만 추출
                    if amount_str:
                        amount_clean = amount_str.replace(',', '').replace('원', '').strip()
                        if amount_clean.isdigit():
                            amount = float(amount_clean)
                            print(f"{company_name} 고지서 금액 조회: {amount:,.0f}원")
                            return amount
                        else:
                            print(f" 금액 형식 오류: {amount_str}")
                            return None
                else:
                    print(f" {company_name} 고지서 정보를 찾을 수 없습니다")
                    return None
            else:
                print(" 고지서 정보 조회 실패")
                return None
        except Exception as e:
            print(f"고지서 금액 조회 실패: {e}")
            return None
    
    def calculate_amount_without_vat(self, total_amount):
        """부가세 10%를 제외한 금액 계산 (1원 오차 보정 포함)"""
        try:
            # 부가세 포함 금액에서 부가세 제외
            # 총금액 = 공급가액 + 부가세(공급가액의 10%)
            # 총금액 = 공급가액 × 1.1
            # 공급가액 = 총금액 / 1.1
            amount_without_vat = round(total_amount / 1.1)
            
            # 검증: 부가세 제외 금액 + 부가세 = 실제 고지서 금액인지 확인
            calculated_total = round(amount_without_vat * 1.1)
            difference = total_amount - calculated_total
            
            if difference != 0:
                # 1원 차이가 있으면 부가세 제외 금액을 보정
                amount_without_vat += difference
                print(f"1원 오차 보정: {difference:+d}원 조정")
                
            # 최종 검증
            final_total = round(amount_without_vat * 1.1)
            print(f"부가세 제외 계산: {total_amount:,.0f}원 → {amount_without_vat}원 (검증: {final_total:,}원)")
            
            return amount_without_vat
        except Exception as e:
            print(f"부가세 제외 계산 실패: {e}")
            return None
    
    def calculate_meta_ics_usage(self, collection_date):
        """Meta ICS 사용량 자동 계산 (SK일렉링크용)"""
        try:
            date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            days_in_month = calendar.monthrange(date_obj.year, date_obj.month)[1]
            
            # SK일렉링크 계정 정보 (고정값)
            metalcs_accounts = 14  # D40 셀의 계정 수
            ssl_vpn_accounts = 0   # SSL-VPN은 사용하지 않음
            
            # 사용일수 계산 (매일 사용 가정)
            metalcs_usage_days = metalcs_accounts * days_in_month
            ssl_vpn_usage_days = ssl_vpn_accounts * days_in_month
            
            # 라이선스 계산
            metalcs_licenses = metalcs_usage_days / days_in_month
            ssl_vpn_licenses = ssl_vpn_usage_days / days_in_month
            
            print(f"SK일렉링크 Meta ICS 사용량 계산:")
            print(f"   - 해당 월 일수: {days_in_month}일")
            print(f"   - MetaLCS 사용일수: {metalcs_usage_days}일 ({metalcs_accounts}계정 × {days_in_month}일)")
            print(f"   - SSL-VPN 사용일수: {ssl_vpn_usage_days}일 ({ssl_vpn_accounts}계정 × {days_in_month}일)")
            print(f"   - MetaLCS 라이선스: {metalcs_licenses}개")
            print(f"   - SSL-VPN 라이선스: {ssl_vpn_licenses}개")
            
            return {
                'days_in_month': days_in_month,
                'metalcs_usage_days': metalcs_usage_days,
                'ssl_vpn_usage_days': ssl_vpn_usage_days,
                'metalcs_licenses': metalcs_licenses,
                'ssl_vpn_licenses': ssl_vpn_licenses
            }
        except Exception as e:
            print(f"Meta ICS 사용량 계산 실패: {e}")
            return None
    
    def update_sk_template(self, template_path, amount_without_vat, collection_date, total_amount):
        """skelectlink.xlsx 템플릿 파일 업데이트"""
        try:
            date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            year_month = f"{date_obj.year}년 {date_obj.month:02d}월"
            date_prefix = f"{str(date_obj.year)[2:]}{date_obj.month:02d}"
            
            # 출력 파일명 생성
            output_filename = f"{date_prefix}_SK일렉링크_청구내역서.xlsx"
            output_path = os.path.join(self.download_dir, output_filename)
            
            # 템플릿 파일 복사
            shutil.copy2(template_path, output_path)
            print(f" 템플릿 파일 복사 완료: {output_filename}")
            
            workbook = load_workbook(output_path)
            
            # B9 셀에 문서번호 설정 (MMP-{년월} 형식)
            document_number = f"MMP-{date_prefix}"
            for sheet in workbook.worksheets:
                if '세부내역' in sheet.title or '대외공문' in sheet.title:
                    sheet.cell(row=9, column=2).value = f"문서번호  : {document_number}"
                    print(f" {sheet.title} B9 셀에 문서번호 설정 완료: {document_number}")
            
            # 1. 대외공문 시트 업데이트 (코오롱과 동일한 로직)
            if '대외공문' in workbook.sheetnames:
                doc_sheet = workbook['대외공문']
                
                # B13 셀 업데이트
                b13_cell = doc_sheet.cell(row=13, column=2)
                if b13_cell.value and isinstance(b13_cell.value, str):
                    old_text = b13_cell.value
                    new_text = re.sub(r'\d{4}년 \d{1,2}월', year_month, old_text)
                    b13_cell.value = new_text
                    print(f" 대외공문 B13 셀 업데이트: {old_text} → {new_text}")
                
                # B16 셀 업데이트 (B,C,D,E,F,G 16행 병합)
                b16_cell = doc_sheet.cell(row=16, column=2)
                if b16_cell.value and isinstance(b16_cell.value, str):
                    old_text = b16_cell.value
                    new_text = re.sub(r'2025년\d{1,2}월', year_month, old_text)
                    if new_text == old_text:
                        new_text = re.sub(r'\d{4}년\d{1,2}월', year_month, old_text)
                    b16_cell.value = new_text
                    print(f" 대외공문 B16 셀 업데이트: {old_text} → {new_text}")
                
                # 하단 테이블 수식 업데이트 (B24, D24, D25)
                self.update_formula_references(doc_sheet, year_month)
            
            # 2. 시트명 변경 및 해당 시트의 B,C,D,E1 병합 셀 텍스트 업데이트
            for sheet in workbook.worksheets:
                if re.match(r'\d{4}년 \d{1,2}월', sheet.title):
                    old_title = sheet.title
                    sheet.title = year_month
                    print(f" 시트명 변경: {old_title} → {year_month}")
                    
                    # B,C,D,E1 병합 셀의 텍스트 업데이트
                    b1_cell = sheet.cell(row=1, column=2)  # B1 셀
                    if b1_cell.value and isinstance(b1_cell.value, str):
                        old_text = b1_cell.value
                        new_text = re.sub(r'\d{4}년 \d{1,2}월', year_month, old_text)
                        if new_text != old_text:
                            b1_cell.value = new_text
                            print(f" {year_month} 시트 B1 셀 텍스트 업데이트: {old_text} → {new_text}")
                    break
            
            # 3. 세부내역 시트 업데이트 - E4 셀에 부가세 제외 금액 입력 (쉼표 없이)
            if '세부내역' in workbook.sheetnames:
                detail_sheet = workbook['세부내역']
                # E4 셀에 부가세 제외 금액 입력 (쉼표 없이)
                detail_sheet.cell(row=4, column=5).value = amount_without_vat
                print(f" 세부내역 시트 E4 셀 업데이트: {amount_without_vat}원 (쉼표 없이)")
                
                # E8 셀에 실제 고지서 청구비용 그대로 입력
                detail_sheet.cell(row=8, column=5).value = total_amount
                print(f" 세부내역 시트 E8 셀 업데이트: {total_amount}원 (실제 고지서 청구비용)")
                
                # Meta ICS 계산 및 업데이트
                ics_data = self.calculate_meta_ics_usage(collection_date)
                if ics_data:
                    # E38 셀에 해당 월 일수 (E40 셀의 =D40*E38 수식이 자동으로 계산됨)
                    detail_sheet.cell(row=38, column=5).value = ics_data['days_in_month']
                    print(f"세부내역 시트 E38 셀 업데이트: {ics_data['days_in_month']}일")
                    
                    print(f"Meta ICS 계산 완료: {ics_data['days_in_month']}일 기준")
            
            # 로고 이미지 삽입 (B2 셀)
            try:
                logo_path = os.path.join(os.path.dirname(__file__), 'assets', 'logo.png')
                if os.path.exists(logo_path):
                    img = Image(logo_path)
                    if '대외공문' in workbook.sheetnames:
                        doc_sheet = workbook['대외공문']
                        doc_sheet.add_image(img, 'B2')
                        print("대외공문 시트 B2 셀에 로고 이미지 삽입 완료")
                    elif workbook.worksheets:
                        workbook.worksheets[0].add_image(img, 'B2')
                        print(f"{workbook.worksheets[0].title} 시트 B2 셀에 로고 이미지 삽입 완료")
            except Exception as e:
                print(f"로고 이미지 삽입 실패: {e}")
            
            # 파일 저장
            workbook.save(output_path)
            workbook.close()
            
            print(f" skelectlink.xlsx 템플릿 업데이트 완료: {output_filename}")
            return output_path
            
        except Exception as e:
            print(f" skelectlink.xlsx 템플릿 업데이트 실패: {e}")
            return None
    
    def update_formula_references(self, doc_sheet, year_month):
        """대외공문 시트의 수식에서 시트명 참조 업데이트"""
        try:
            # 수식이 있을 수 있는 셀들 확인 (24행, 25행 등)
            cells_to_check = [
                (24, 2), (24, 3), (24, 4),  # B24, C24, D24
                (25, 2), (25, 3), (25, 4),  # B25, C25, D25
            ]
            
            for row, col in cells_to_check:
                cell = doc_sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    old_formula = cell.value
                    # 수식에서 시트명 참조 패턴 찾아서 교체
                    new_formula = re.sub(r"'(\d{4}년 \d{1,2}월)'!", f"'{year_month}'!", old_formula)
                    if new_formula != old_formula:
                        cell.value = new_formula
                        print(f" 대외공문 {chr(64+col)}{row} 셀 수식 업데이트: {old_formula} → {new_formula}")
                        
        except Exception as e:
            print(f"수식 참조 업데이트 오류: {e}")
    
    def process_sk_data(self, collection_date):
        """SK일렉링크 데이터 전처리 메인 함수"""
        try:
            print(" SK일렉링크 데이터 전처리 시작")
            
            # 1. 고지서 금액 조회
            total_amount = self.get_bill_amount("SK일렉링크")
            if total_amount is None:
                print(" SK일렉링크 고지서 금액을 찾을 수 없습니다")
                return False
            
            # 2. 부가세 제외 금액 계산
            amount_without_vat = self.calculate_amount_without_vat(total_amount)
            if amount_without_vat is None:
                print(" 부가세 제외 계산 실패")
                return False
            
            # 3. skelectlink.xlsx 템플릿 다운로드
            template_path = self.download_sk_template()
            if template_path is None:
                print(" skelectlink.xlsx 템플릿 다운로드 실패")
                return False
            
            # 4. 템플릿 업데이트 및 청구서 생성
            final_invoice_path = self.update_sk_template(template_path, amount_without_vat, collection_date, total_amount)
            if final_invoice_path is None:
                print("SK일렉링크 청구서 생성 실패")
                return False
            
            print(f"SK일렉링크 전처리 완료! 파일 생성:")
            print(f"   - SK일렉링크 청구내역서: {os.path.basename(final_invoice_path)}")
            
            return True
            
        except Exception as e:
            import traceback
            print(f"SK일렉링크 전처리 실패: {e}")
            print(f"상세 에러: {traceback.format_exc()}")
            return False
