"""청구서 템플릿 공통: 대외공문 시트 대표이사 표기 등."""

from __future__ import annotations

from typing import Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 템플릿에서 흔히 쓰는 형태: "대표이사" + 공백 3칸 + 이름(이름 안에 띄어쓰기 허용)
CEO_PREFIX = "대표이사"
CEO_SEPARATOR = "   "


def _write_cell_respecting_merge(sheet, row: int, col: int, value) -> Tuple[int, int]:
    """
    엑셀 병합 셀은 openpyxl에서 좌상단만 쓰기 가능.
    반환: 실제로 값을 쓴 (row, col).
    """
    for mrange in sheet.merged_cells.ranges:
        if mrange.min_row <= row <= mrange.max_row and mrange.min_col <= col <= mrange.max_col:
            sheet.cell(row=mrange.min_row, column=mrange.min_col).value = value
            return (mrange.min_row, mrange.min_col)
    sheet.cell(row=row, column=col).value = value
    return (row, col)


def format_ceo_line_value(ceo_name: Optional[str]) -> Optional[str]:
    """저장된 이름으로 대외공문용 한 줄 문자열 생성. 이름이 비어 있으면 None(셀 건드리지 않음)."""
    if ceo_name is None:
        return None
    name = str(ceo_name).strip()
    if not name:
        return None
    return f"{CEO_PREFIX}{CEO_SEPARATOR}{name}"


def apply_ceo_line_to_doc_sheet_d35(
    workbook: Workbook,
    ceo_name: Optional[str] = None,
    sheet_name: str = "대외공문",
    row: int = 35,
    col: int = 4,
) -> bool:
    """
    대외공문 시트 D35(기본)에 '대표이사   {이름}' 형식으로 기록.
    병합 셀이면 해당 범위 좌상단에 기록.
    ceo_name이 None이면 AdminStorage에서 읽음.
    """
    if sheet_name not in workbook.sheetnames:
        print(f"⚠️ '{sheet_name}' 시트 없음 — 대표이사 줄 생략")
        return False

    if ceo_name is None:
        from backend.storage.admin_storage import AdminStorage

        data = AdminStorage().get_invoice_common_settings()
        ceo_name = data.get("ceo_name")

    val = format_ceo_line_value(ceo_name)
    if val is None:
        return False

    doc_sheet = workbook[sheet_name]
    r, c = _write_cell_respecting_merge(doc_sheet, row, col, val)
    addr = f"{get_column_letter(c)}{r}"
    note = ""
    if (r, c) != (row, col):
        note = f" (템플릿 병합: {get_column_letter(col)}{row} → 실제 기록 {addr})"
    print(f"{sheet_name} 시트 {addr} 셀 업데이트 (대표이사){note}: {val}")
    return True
