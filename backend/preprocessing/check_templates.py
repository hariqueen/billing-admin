import os, json, firebase_admin
from dotenv import load_dotenv
from firebase_admin import credentials, storage
from openpyxl import load_workbook

# Firebase 초기화
load_dotenv()
cred_dict = json.loads(os.environ["FIREBASE_PRIVATE_KEY"])
BUCKET_NAME = os.getenv("STORAGE_BUCKET", "services-e42af.firebasestorage.app")

if not firebase_admin._apps:
    firebase_admin.initialize_app(
        credentials.Certificate(cred_dict),
        {"storageBucket": BUCKET_NAME}
    )

bucket = storage.bucket()
template_files = ['annhouse.xlsx', 'annhouse_CS.xlsx', 'annhouse_TS.xlsx']

for template_name in template_files:
    try:
        print(f"\n📋 템플릿 파일 확인: {template_name}")
        
        # 파일 다운로드
        blob = bucket.blob(template_name)
        temp_path = f"temp_{template_name}"
        blob.download_to_filename(temp_path)
        
        # 워크북 구조 확인
        workbook = load_workbook(temp_path)
        print(f"   시트 목록: {workbook.sheetnames}")
        
        # 각 시트의 행/열 수 확인
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"   - {sheet_name}: {sheet.max_row}행 x {sheet.max_column}열")
        
        workbook.close()
        
        # 임시 파일 삭제
        os.remove(temp_path)
        
    except Exception as e:
        print(f"❌ {template_name} 확인 실패: {e}")
