import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import firebase_admin
from firebase_admin import credentials, storage
from backend.utils.secrets_manager import get_firebase_secret

class WconceptPreprocessor:
    def __init__(self):
        self.download_dir = os.path.join(os.getcwd(), "temp_processing")
        os.makedirs(self.download_dir, exist_ok=True)
        self.setup_firebase()
    
    def setup_firebase(self):
        """Firebase Storage 연결 설정"""
        try:
            from dotenv import load_dotenv
            
            load_dotenv()
            cred_dict = get_firebase_secret()
            
            BUCKET_NAME = os.getenv("STORAGE_BUCKET", "services-e42af.firebasestorage.app")
            
            # 기존 앱이 있으면 삭제하고 새로 초기화
            if firebase_admin._apps:
                firebase_admin.delete_app(firebase_admin.get_app())
            
            firebase_admin.initialize_app(
                credentials.Certificate(cred_dict),
                {"storageBucket": BUCKET_NAME}
            )
            
            self.bucket = storage.bucket()
            print("Firebase Storage 연결 완료")
        except Exception as e:
            print(f"Firebase 연결 실패: {e}")
            self.bucket = None
    
    def download_wconcept_template(self):
        """Firebase에서 wconcept.xlsx 템플릿 다운로드"""
        if not self.bucket:
            print("Firebase 연결이 없습니다")
            return None
        
        try:
            blob = self.bucket.blob("wconcept.xlsx")
            temp_dir = os.path.join(os.getcwd(), "temp_processing")
            os.makedirs(temp_dir, exist_ok=True)
            
            local_path = os.path.join(temp_dir, "wconcept.xlsx")
            blob.download_to_filename(local_path)
            print(f"wconcept.xlsx 템플릿 다운로드 완료: {local_path}")
            return local_path
        except Exception as e:
            print(f"wconcept.xlsx 템플릿 다운로드 실패: {e}")
            return None
    
    def update_wconcept_template(self, template_path, license_count, collection_date, bill_amount=None):
        """wconcept.xlsx 템플릿 파일 업데이트 (W컨셉은 전달 청구서)"""
        try:
            date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            
            # W컨셉은 n-1월로 처리 (전달 청구서)
            if date_obj.month == 1:
                prev_year = date_obj.year - 1
                prev_month = 12
            else:
                prev_year = date_obj.year
                prev_month = date_obj.month - 1
            
            year_month = f"{prev_year}년 {prev_month:02d}월"
            date_prefix = f"{str(prev_year)[2:]}{prev_month:02d}"
            
            # 출력 파일명 생성
            output_filename = f"{date_prefix}_W컨셉_청구내역서.xlsx"
            output_path = os.path.join(self.download_dir, output_filename)
            
            # 템플릿 파일 복사
            shutil.copy2(template_path, output_path)
            print(f"템플릿 파일 복사 완료: {output_filename}")
            
            workbook = load_workbook(output_path)
            
            # B9 셀에 문서번호 설정 (MMP-{년월} 형식)
            document_number = f"MMP-{date_prefix}"
            for sheet in workbook.worksheets:
                if '세부내역' in sheet.title or '대외공문' in sheet.title:
                    sheet.cell(row=9, column=2).value = f"문서번호  : {document_number}"
                    print(f" {sheet.title} B9 셀에 문서번호 설정 완료: {document_number}")
            
            # 1. 대외공문 시트 업데이트 (SK일렉링크와 동일한 로직)
            if '대외공문' in workbook.sheetnames:
                doc_sheet = workbook['대외공문']
                
                # B13 셀 업데이트
                b13_cell = doc_sheet.cell(row=13, column=2)
                if b13_cell.value and isinstance(b13_cell.value, str):
                    old_text = b13_cell.value
                    new_text = re.sub(r'\d{4}년 \d{1,2}월', year_month, old_text)
                    b13_cell.value = new_text
                    print(f" 대외공문 B13 셀 업데이트: {old_text} → {new_text}")
                
                # B16 셀 업데이트 (B,C,D,E,F,G 16행 병합)
                b16_cell = doc_sheet.cell(row=16, column=2)
                if b16_cell.value and isinstance(b16_cell.value, str):
                    old_text = b16_cell.value
                    new_text = re.sub(r'2025년\d{1,2}월', year_month, old_text)
                    if new_text == old_text:
                        new_text = re.sub(r'\d{4}년\d{1,2}월', year_month, old_text)
                    b16_cell.value = new_text
                    print(f"대외공문 B16 셀 업데이트: {old_text} → {new_text}")
                
                # 하단 테이블 수식 업데이트 (B24, D24, D25)
                self.update_formula_references(doc_sheet, year_month)
            
            # 2. 시트명 변경 및 해당 시트의 B,C,D,E1 병합 셀 텍스트 업데이트
            for sheet in workbook.worksheets:
                if re.match(r'\d{4}년 \d{1,2}월', sheet.title):
                    old_title = sheet.title
                    sheet.title = year_month
                    print(f"시트명 변경: {old_title} → {year_month}")
                    
                    # B,C,D,E1 병합 셀의 텍스트 업데이트
                    b1_cell = sheet.cell(row=1, column=2)  # B1 셀
                    if b1_cell.value and isinstance(b1_cell.value, str):
                        old_text = b1_cell.value
                        new_text = re.sub(r'\d{4}년 \d{1,2}월', year_month, old_text)
                        if new_text != old_text:
                            b1_cell.value = new_text
                            print(f"{year_month} 시트 B1 셀 텍스트 업데이트: {old_text} → {new_text}")
                    break
            
            # 3. 세부내역 시트 업데이트
            if '세부내역' in workbook.sheetnames:
                detail_sheet = workbook['세부내역']
                # D5 셀에 라이선스 수량 입력 (숫자만)
                detail_sheet.cell(row=5, column=4).value = license_count
                print(f"세부내역 시트 D5 셀 업데이트: {license_count}개")
                
                # E17 셀에 고지서 금액에서 부가세 10% 제외한 금액 입력
                if bill_amount:
                    # 고지서 금액에서 숫자만 추출 (예: "862,120원" → 862120)
                    amount_str = re.sub(r'[^\d]', '', bill_amount)
                    if amount_str:
                        total_amount = int(amount_str)
                        # 부가세 10% 제외한 금액 계산 (1원 오차 보정 포함)
                        net_amount = round(total_amount / 1.1)
                        
                        # 검증: 부가세 제외 금액 + 부가세 = 실제 고지서 금액인지 확인
                        calculated_total = round(net_amount * 1.1)
                        difference = total_amount - calculated_total
                        
                        if difference != 0:
                            # 1원 차이가 있으면 부가세 제외 금액을 보정
                            net_amount += difference
                            print(f"1원 오차 보정: {difference:+d}원 조정")
                            
                        # 최종 검증
                        final_total = round(net_amount * 1.1)
                        
                        detail_sheet.cell(row=17, column=5).value = net_amount
                        print(f"세부내역 시트 E17 셀 업데이트: {net_amount:,}원 (고지서: {total_amount:,}원, 검증: {final_total:,}원)")
                        
                        # E21 셀에 실제 고지서 청구비용 그대로 입력
                        detail_sheet.cell(row=21, column=5).value = total_amount
                        print(f"세부내역 시트 E21 셀 업데이트: {total_amount:,}원 (실제 고지서 청구비용)")
            
            # 로고 이미지 삽입 (B2 셀)
            try:
                logo_path = os.path.join(os.path.dirname(__file__), 'assets', 'logo.png')
                if os.path.exists(logo_path):
                    img = Image(logo_path)
                    if '대외공문' in workbook.sheetnames:
                        doc_sheet = workbook['대외공문']
                        doc_sheet.add_image(img, 'B2')
                        print("대외공문 시트 B2 셀에 로고 이미지 삽입 완료")
                    elif workbook.worksheets:
                        workbook.worksheets[0].add_image(img, 'B2')
                        print(f"{workbook.worksheets[0].title} 시트 B2 셀에 로고 이미지 삽입 완료")
            except Exception as e:
                print(f"로고 이미지 삽입 실패: {e}")
            
            # 파일 저장
            workbook.save(output_path)
            workbook.close()
            
            print(f"wconcept.xlsx 템플릿 업데이트 완료: {output_filename}")
            return output_path
            
        except Exception as e:
            print(f"wconcept.xlsx 템플릿 업데이트 실패: {e}")
            return None
    
    def update_formula_references(self, doc_sheet, year_month):
        """대외공문 시트의 수식에서 시트명 참조 업데이트"""
        try:
            # 수식이 있을 수 있는 셀들 확인 (24행, 25행 등)
            cells_to_check = [
                (24, 2), (24, 3), (24, 4),  # B24, C24, D24
                (25, 2), (25, 3), (25, 4),  # B25, C25, D25
            ]
            
            for row, col in cells_to_check:
                cell = doc_sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    old_formula = cell.value
                    # 수식에서 시트명 참조 패턴 찾아서 교체
                    new_formula = re.sub(r"'(\d{4}년 \d{1,2}월)'!", f"'{year_month}'!", old_formula)
                    if new_formula != old_formula:
                        cell.value = new_formula
                        print(f"대외공문 {chr(64+col)}{row} 셀 수식 업데이트: {old_formula} → {new_formula}")
                        
        except Exception as e:
            print(f"수식 참조 업데이트 오류: {e}")
    
    def process_wconcept_data(self, collection_date, license_count=40):
        """W컨셉 데이터 전처리 메인 함수"""
        try:
            print(" W컨셉 데이터 전처리 시작")
            print(f" 청구 라이선스 수량: {license_count}개")
            
            # 고지서 금액 조회
            from ..storage.admin_storage import AdminStorage
            admin_storage = AdminStorage()
            bill_amounts = admin_storage.get_bill_amounts()
            bill_amount = bill_amounts.get("W컨셉", {}).get("amount")
            
            if bill_amount:
                print(f"W컨셉 고지서 금액: {bill_amount}")
            else:
                print("W컨셉 고지서 금액을 찾을 수 없습니다")
            
            # 1. wconcept.xlsx 템플릿 다운로드
            template_path = self.download_wconcept_template()
            if template_path is None:
                print("wconcept.xlsx 템플릿 다운로드 실패")
                return False
            
            # 2. 템플릿 업데이트 및 청구서 생성 (고지서 금액 포함)
            final_invoice_path = self.update_wconcept_template(template_path, license_count, collection_date, bill_amount)
            if final_invoice_path is None:
                print("W컨셉 청구서 생성 실패")
                return False
            
            print(f"W컨셉 전처리 완료! 파일 생성:")
            print(f"   - W컨셉 청구내역서: {os.path.basename(final_invoice_path)}")
            
            return True
            
        except Exception as e:
            import traceback
            print(f" W컨셉 전처리 실패: {e}")
            print(f"상세 에러: {traceback.format_exc()}")
            return False
