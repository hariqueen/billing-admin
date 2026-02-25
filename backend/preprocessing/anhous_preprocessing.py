import os
import re
import shutil
import pandas as pd
from datetime import datetime
from pathlib import Path
from io import StringIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import firebase_admin
from firebase_admin import credentials, storage
from backend.utils.secrets_manager import get_firebase_secret
import calendar

class AnhousPreprocessor:
    def __init__(self):
        self.download_dir = os.path.join(os.getcwd(), "temp_processing")
        os.makedirs(self.download_dir, exist_ok=True)
        self.setup_firebase()
    
    def setup_firebase(self):
        """Firebase Storage ì—°ê²° ì„¤ì •"""
        try:
            cred_dict = get_firebase_secret()
            BUCKET_NAME = "services-e42af.firebasestorage.app"
            
            # ê¸°ì¡´ ì•±ì´ ìˆìœ¼ë©´ ì‚­ì œí•˜ê³  ìƒˆë¡œ ì´ˆê¸°í™”
            if firebase_admin._apps:
                firebase_admin.delete_app(firebase_admin.get_app())
            
            firebase_admin.initialize_app(
                credentials.Certificate(cred_dict),
                {"storageBucket": BUCKET_NAME}
            )
            
            self.bucket = storage.bucket()
            print(" Firebase Storage ì—°ê²° ì™„ë£Œ")
        except Exception as e:
            print(f" Firebase ì—°ê²° ì‹¤íŒ¨: {e}")
            self.bucket = None
    
    def convert_xls_to_csv(self, xls_file_path):
        """XLS/XLSX íŒŒì¼ì„ CSVë¡œ ë³€í™˜"""
        try:
            print(f"ğŸ” ë³€í™˜ ì‹œì‘: {os.path.basename(xls_file_path)}")
            print(f"   ì›ë³¸ íŒŒì¼: {xls_file_path}")
            
            # íŒŒì¼ í™•ì¥ì í™•ì¸
            if xls_file_path.endswith('.xlsx'):
                # Excel íŒŒì¼ ì²˜ë¦¬
                print(f"   Excel íŒŒì¼ë¡œ ì½ê¸° ì‹œë„...")
                df = pd.read_excel(xls_file_path)
                print(f"   ë°ì´í„° í˜•íƒœ: {df.shape[0]}í–‰ x {df.shape[1]}ì—´")
                print(f"   ì»¬ëŸ¼: {list(df.columns)}")
                csv_path = xls_file_path.replace('.xlsx', '.csv')
                df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                print(f" Excel â†’ .csv ë³€í™˜ ì™„ë£Œ: {csv_path}")
                return csv_path
            else:
                # XLS íŒŒì¼ ì²˜ë¦¬ (HTML ë‚´ìš©ì¸ì§€ í™•ì¸)
                try:
                    with open(xls_file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    # HTML ë‚´ìš©ì¸ì§€ í™•ì¸
                    if '<html' in content.lower() or '<table' in content.lower():
                        # HTML í…Œì´ë¸” íŒŒì‹±
                        df = pd.read_html(StringIO(content))[0]
                        
                        # ì²« ë²ˆì§¸ í–‰ì´ ì œëª©ì¸ ê²½ìš° ì œê±°
                        if df.shape[0] > 0 and 'í†µí™”ë‚´ì—­' in str(df.iloc[0, 0]):
                            df = df.iloc[1:]
                        
                        # í—¤ë” ì„¤ì •
                        if df.shape[0] > 0:
                            df.columns = df.iloc[0]
                            df = df.iloc[1:]
                        
                        # CSVë¡œ ì €ì¥
                        csv_path = xls_file_path.replace('.xls', '.csv')
                        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                        print(f" HTML â†’ .csv ë³€í™˜ ì™„ë£Œ: {csv_path}")
                        return csv_path
                    else:
                        # ì¼ë°˜ Excel íŒŒì¼ ì²˜ë¦¬
                        df = pd.read_excel(xls_file_path)
                        csv_path = xls_file_path.replace('.xls', '.csv')
                        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                        print(f" Excel â†’ .csv ë³€í™˜ ì™„ë£Œ: {csv_path}")
                        return csv_path
                except UnicodeDecodeError:
                    # ì¸ì½”ë”© ì˜¤ë¥˜ ì‹œ Excelë¡œ ì½ê¸° ì‹œë„
                    df = pd.read_excel(xls_file_path)
                    csv_path = xls_file_path.replace('.xls', '.csv')
                    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                    print(f" Excel â†’ .csv ë³€í™˜ ì™„ë£Œ: {csv_path}")
                    return csv_path
                
        except Exception as e:
            print(f" ë³€í™˜ ì‹¤íŒ¨: {xls_file_path}, ì˜¤ë¥˜: {e}")
            return None
    
    def convert_all_collected_files(self):
        """ì—…ë¡œë“œëœ íŒŒì¼ë“¤ë§Œ CSVë¡œ ë³€í™˜"""
        csv_files = []
        
        temp_dir = "temp_processing"
        if os.path.exists(temp_dir):
            anhous_files = []
            for filename in os.listdir(temp_dir):
                if "ì•¤í•˜ìš°ìŠ¤" in filename and filename.endswith((".xls", ".xlsx")):
                    file_path = os.path.join(temp_dir, filename)
                    anhous_files.append((file_path, os.path.getmtime(file_path)))
            
            # ìµœì‹  íŒŒì¼ 2ê°œë§Œ ì„ íƒ (SMS, CALL)
            anhous_files.sort(key=lambda x: x[1], reverse=True)
            latest_files = anhous_files[:2]
            
            for file_path, _ in latest_files:
                csv_path = self.convert_xls_to_csv(file_path)
                if csv_path:
                    csv_files.append(csv_path)
        
        return csv_files
    
    def find_anhous_templates(self):
        """Firebase Storageì—ì„œ ì•¤í•˜ìš°ìŠ¤ í…œí”Œë¦¿ íŒŒì¼ë“¤ ì°¾ê¸°"""
        if not self.bucket:
            return []
        
        templates = []
        blobs = self.bucket.list_blobs()
        
        for blob in blobs:
            if 'annhouse' in blob.name.lower():
                templates.append(blob.name)
                print(f" í…œí”Œë¦¿ ë°œê²¬: {blob.name}")
        
        return templates
    
    def download_template(self, template_name):
        """í…œí”Œë¦¿ íŒŒì¼ ë‹¤ìš´ë¡œë“œ"""
        if not self.bucket:
            return None
        
        try:
            blob = self.bucket.blob(template_name)
            temp_dir = os.path.join(os.getcwd(), "temp_processing")
            os.makedirs(temp_dir, exist_ok=True)
            
            local_path = os.path.join(temp_dir, template_name)
            blob.download_to_filename(local_path)
            print(f" {template_name} í…œí”Œë¦¿ ë‹¤ìš´ë¡œë“œ ì™„ë£Œ: {local_path}")
            return local_path
        except Exception as e:
            print(f" í…œí”Œë¦¿ ë‹¤ìš´ë¡œë“œ ì‹¤íŒ¨: {e}")
            return None
    
    def get_call_data_by_team(self, csv_files, collection_date):
        """CSV íŒŒì¼ë“¤ì—ì„œ íŒ€ë³„ CALL ë°ì´í„° ë¶„ë¥˜"""
        team_data = {}
        
        for csv_file in csv_files:
            try:
                # ê³ ê°ë²ˆí˜¸ ì»¬ëŸ¼ì„ ë¬¸ìì—´ë¡œ ì½ì–´ì„œ ì•ì˜ 0ì´ ì‚¬ë¼ì§€ì§€ ì•Šë„ë¡ í•¨
                df = pd.read_csv(csv_file, dtype={'ê³ ê°ë²ˆí˜¸': str})
                
                if 'íŒ€' in list(df.columns):
                    teams = df['íŒ€'].unique()
                    print(f" CSV íŒŒì¼ ë¶„ì„: {csv_file}")
                    print(f"   íŒ€ ì¢…ë¥˜: {teams}")
                    
                    for team in teams:
                        team_df = df[df['íŒ€'] == team].copy()
                        if team not in team_data:
                            team_data[team] = []
                        team_data[team].append(team_df)
                        print(f"    {team} ë°ì´í„°: {len(team_df)}í–‰")
                        
            except Exception as e:
                print(f" CSV íŒŒì¼ ì½ê¸° ì‹¤íŒ¨: {csv_file}, ì˜¤ë¥˜: {e}")
        
        return team_data
    
    def update_template_with_team_data(self, template_path, team_data, collection_date, template_team, csv_files):
        """í…œí”Œë¦¿ íŒŒì¼ì— íŒ€ë³„ ë°ì´í„° ì—…ë°ì´íŠ¸ (ì›ë³¸ í…œí”Œë¦¿ ìœ ì§€)"""
        try:
            import shutil
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image
            
            print(f"{template_team} ë°ì´í„° ì²˜ë¦¬ ì¤‘...")
            
            # ì›ë³¸ í…œí”Œë¦¿ íŒŒì¼ ë³µì‚¬
            date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            date_prefix = f"{str(date_obj.year)[2:]}{date_obj.month:02d}"
            
            base_name = os.path.basename(template_path)
            if 'annhouse_CS' in base_name:
                team_suffix = 'CS'
            elif 'annhouse_TS' in base_name:
                team_suffix = 'TS'
            elif base_name == 'annhouse.xlsx':
                team_suffix = 'ì°½ì—…'
            else:
                team_suffix = 'unknown'
            
            new_filename = f"{date_prefix}_ì•¤í•˜ìš°ìŠ¤ ìˆ˜ìˆ˜ë£Œ ì²­êµ¬ë‚´ì—­ì„œ_{team_suffix}.xlsx"
            
            # ë‹¤ìš´ë¡œë“œ í´ë”ì— ì €ì¥ (/app/downloadsë¡œ í†µì¼)
            download_dir = "/app/downloads"
            os.makedirs(download_dir, exist_ok=True)
            output_path = os.path.join(download_dir, new_filename)
            
            # ì›ë³¸ í…œí”Œë¦¿ ë³µì‚¬
            shutil.copy2(template_path, output_path)
            
            workbook = load_workbook(output_path)
            
            # í†µí™”ë£Œ ì‹œíŠ¸ ê°€ì ¸ì˜¤ê¸°
            if 'í†µí™”ë£Œ' in workbook.sheetnames:
                sheet = workbook['í†µí™”ë£Œ']
            else:
                print("í†µí™”ë£Œ ì‹œíŠ¸ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
                return None
            
            # í•´ë‹¹ íŒ€ì˜ ë°ì´í„°ë§Œ ì²˜ë¦¬
            if template_team not in team_data:
                print(f"{template_team} ë°ì´í„°ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
                return None
            
            team_dataframes = team_data[template_team]
            if not team_dataframes:
                print(f"{template_team} ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤")
                return None
            
            # DataFrame ê°ì²´ë¥¼ ì§ì ‘ ì‚¬ìš© (CSV íŒŒì¼ ê²½ë¡œê°€ ì•„ë‹˜)
            df = team_dataframes[0]  # ì²« ë²ˆì§¸ DataFrame
            
            # ë‚ ì§œ ì •ë³´ ì¶”ì¶œ
            year_month = f"{date_obj.year}ë…„ {date_obj.month}ì›”"
            
            # H3 ì…€ì— ê³¼ê¸ˆì›” ì •ë³´ ì¶”ê°€
            print(f"   H3 ì…€ì— ê³¼ê¸ˆì›” ì„¤ì • ì‹œë„: {year_month}")
            print(f"   í˜„ì¬ ì‹œíŠ¸: {sheet.title}")
            print(f"   H3 ì…€ ìœ„ì¹˜: row=3, column=8")
            
            sheet.cell(row=3, column=8).value = year_month  # H3 ì…€
            
            # ì„¤ì • í™•ì¸
            h3_value = sheet.cell(row=3, column=8).value
            print(f"H3 ì…€ì— ê³¼ê¸ˆì›” ì„¤ì • ì™„ë£Œ: {h3_value}")
            
            # B9 ì…€ì— ë¬¸ì„œë²ˆí˜¸ ì„¤ì • (MMP-{ë…„ì›”} í˜•ì‹)
            document_number = f"MMP-{date_prefix}"
            sheet.cell(row=9, column=2).value = f"ë¬¸ì„œë²ˆí˜¸  : {document_number}"
            print(f"B9 ì…€ì— ë¬¸ì„œë²ˆí˜¸ ì„¤ì • ì™„ë£Œ: {document_number}")
            
            # ì‹œíŠ¸ëª…ê³¼ í…ìŠ¤íŠ¸ ë‚ ì§œ ì—…ë°ì´íŠ¸
            print(f"ì‹œíŠ¸ëª… ë° ë‚ ì§œ í…ìŠ¤íŠ¸ ì—…ë°ì´íŠ¸ ì‹œì‘")
            self.update_sheet_dates(workbook, year_month)
            print(f"ì‹œíŠ¸ëª… ë° ë‚ ì§œ í…ìŠ¤íŠ¸ ì—…ë°ì´íŠ¸ ì™„ë£Œ")
            
            # SMS ë°ì´í„° ì²˜ë¦¬ ë° ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ ì—…ë°ì´íŠ¸
            print(f"SMS ë°ì´í„° ì²˜ë¦¬ ì‹œì‘")
            self.process_sms_data(workbook, template_team, csv_files)
            print(f"SMS ë°ì´í„° ì²˜ë¦¬ ì™„ë£Œ")
            
            # ê¸°ì¡´ ë°ì´í„° ì§€ìš°ê¸° (8í–‰ë¶€í„°)
            for row in range(8, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).value = None
            
            # ë°ì´í„° ì²˜ë¦¬
            current_row = 8
            
            column_mapping = {
                'í†µí™”ì‹œê°„': 'C',
                'ì½œì‹œì‘ì‹œê°„': 'D', 
                'ëŒ€ê¸°ì‹œì‘ì‹œê°„': 'E',
                'ë§ì‹œì‘ì‹œê°„': 'F',
                'í†µí™”ì‹œì‘ì‹œê°„': 'G',
                'ì½œì¢…ë£Œì‹œê°„': 'H'
            }
            
            # ê³„ì‚° í•¨ìˆ˜ë“¤ ì •ì˜
            import math
            
            def calculate_settlement_type(customer_number):
                """ì •ì‚°êµ¬ë¶„ ê³„ì‚°: 010ìœ¼ë¡œ ì‹œì‘í•˜ë©´ ì´ë™ì „í™”, ì•„ë‹ˆë©´ ì‹œë‚´/ì‹œì™¸"""
                return "ì´ë™ì „í™”" if str(customer_number).strip().startswith('010') else "ì‹œë‚´/ì‹œì™¸"
            
            def calculate_seconds(call_time):
                """ì´ˆë‹¨ìœ„ í™˜ì‚°: HH:MM:SS í˜•ì‹ì„ ì´ˆë¡œ ë³€í™˜"""
                try:
                    time_str = str(call_time)
                    if len(time_str) >= 8 and time_str.count(':') == 2:
                        hours, minutes, seconds = map(int, time_str.split(':'))
                        return hours * 3600 + minutes * 60 + seconds
                    return 0
                except:
                    return 0
            
            def calculate_billing_units(settlement_type, total_seconds):
                """ê³¼ê¸ˆêµ¬ê°„ ê³„ì‚°"""
                divisor = 10 if settlement_type == "ì´ë™ì „í™”" else 180
                return math.ceil(total_seconds / divisor)
            
            def calculate_billing_amount(settlement_type, billing_units):
                """ê³¼ê¸ˆê¸ˆì•¡ ê³„ì‚°"""
                rate = 10 if settlement_type == "ì´ë™ì „í™”" else 30
                return billing_units * rate
            
            for idx, call_row in df.iterrows():
                # ê¸°ì¤€ì›” ì¶”ê°€ (Bì—´)
                sheet.cell(row=current_row, column=2).value = year_month
                
                # CALL ë°ì´í„° ë§¤í•‘
                for call_col, excel_col in column_mapping.items():
                    if call_col in list(df.columns):
                        try:
                            col_letter = excel_col
                            col_idx = ord(col_letter) - ord('A') + 1
                            sheet.cell(row=current_row, column=col_idx).value = call_row[call_col]
                        except Exception as e:
                            print(f"ë°ì´í„° ë§¤í•‘ ì˜¤ë¥˜: {e}")
                
                # ì¶”ê°€ ê³„ì‚° í•„ë“œë“¤
                try:
                    # ê³ ê°ë²ˆí˜¸ì™€ í†µí™”ì‹œê°„ ê°€ì ¸ì˜¤ê¸°
                    customer_number = call_row.get('ê³ ê°ë²ˆí˜¸', '')
                    call_time = call_row.get('í†µí™”ì‹œê°„', '00:00:00')
                    
                    # ê³„ì‚° ìˆ˜í–‰
                    settlement_type = calculate_settlement_type(customer_number)
                    total_seconds = calculate_seconds(call_time)
                    billing_units = calculate_billing_units(settlement_type, total_seconds)
                    billing_amount = calculate_billing_amount(settlement_type, billing_units)
                    
                    # ì—‘ì…€ì— ê°’ ì…ë ¥
                    sheet.cell(row=current_row, column=9).value = settlement_type      # Iì—´: ì •ì‚°êµ¬ë¶„
                    sheet.cell(row=current_row, column=10).value = total_seconds      # Jì—´: ì´ˆë‹¨ìœ„ í™˜ì‚°
                    sheet.cell(row=current_row, column=11).value = billing_units      # Kì—´: ê³¼ê¸ˆêµ¬ê°„
                    sheet.cell(row=current_row, column=12).value = billing_amount     # Lì—´: ê³¼ê¸ˆê¸ˆì•¡
                    
                except Exception as e:
                    print(f"ê³„ì‚° ì˜¤ë¥˜ (í–‰ {current_row}): {e}")
                
                current_row += 1
            
            # AJ29 ì…€ì— í•´ë‹¹ ì›” ì¼ìˆ˜ ì—…ë°ì´íŠ¸ (Meta ICS ê³„ì‚°ìš©)
            date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            days_in_month = calendar.monthrange(date_obj.year, date_obj.month)[1]
            sheet.cell(row=29, column=36).value = days_in_month  # AJ29 ì…€
            print(f"ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ AJ29 ì…€ ì—…ë°ì´íŠ¸: {days_in_month}ì¼")
            
            # íŒ€ë³„ ICS ì‚¬ìš© í˜„í™© ì„¤ì •
            aj_col = 36  # AJì—´ = 36ë²ˆì§¸ ì»¬ëŸ¼
            ak_col = 37  # AKì—´ = 37ë²ˆì§¸ ì»¬ëŸ¼
            
            # íŒ€ë³„ í–‰ ë²”ìœ„ ì •ì˜
            team_ranges = {
                "ì°½ì—…": range(31, 39),  # 31-38í–‰
                "TS": range(31, 33),    # 31-32í–‰  
                "CS": range(33, 36)     # 33-35í–‰
            }
            
            # AJ ì—´: í•­ìƒ 1ë¡œ ì„¤ì •
            for team_name, row_range in team_ranges.items():
                for row in row_range:
                    sheet.cell(row=row, column=aj_col).value = 1
                print(f"ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ AJ{min(row_range)}-{max(row_range)} ì…€ ì—…ë°ì´íŠ¸ ({team_name}): 1")
            
            # AK ì—´: 31ì¼ê¹Œì§€ ìˆëŠ” ì›”ì´ë©´ 1, ì•„ë‹ˆë©´ 0
            ak_value = 1 if days_in_month >= 31 else 0
            for team_name, row_range in team_ranges.items():
                for row in row_range:
                    sheet.cell(row=row, column=ak_col).value = ak_value
                print(f"ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ AK{min(row_range)}-{max(row_range)} ì…€ ì—…ë°ì´íŠ¸ ({team_name}): {ak_value} ({'31ì¼ ì›”' if ak_value == 1 else '30ì¼ ì´í•˜ ì›”'})")
            
            # ë¡œê³  ì´ë¯¸ì§€ ì‚½ì… (B2 ì…€)
            try:
                logo_path = os.path.join(os.path.dirname(__file__), 'assets', 'logo.png')
                if os.path.exists(logo_path):
                    img = Image(logo_path)
                    if 'ëŒ€ì™¸ê³µë¬¸' in workbook.sheetnames:
                        doc_sheet = workbook['ëŒ€ì™¸ê³µë¬¸']
                        doc_sheet.add_image(img, 'B2')
                        print("ëŒ€ì™¸ê³µë¬¸ ì‹œíŠ¸ B2 ì…€ì— ë¡œê³  ì´ë¯¸ì§€ ì‚½ì… ì™„ë£Œ")
                    elif workbook.worksheets:
                        workbook.worksheets[0].add_image(img, 'B2')
                        print(f"{workbook.worksheets[0].title} ì‹œíŠ¸ B2 ì…€ì— ë¡œê³  ì´ë¯¸ì§€ ì‚½ì… ì™„ë£Œ")
            except Exception as e:
                print(f"ë¡œê³  ì´ë¯¸ì§€ ì‚½ì… ì‹¤íŒ¨: {e}")
            
            # íŒŒì¼ ì €ì¥
            workbook.save(output_path)
            workbook.close()
            
            print(f" í…œí”Œë¦¿ ì—…ë°ì´íŠ¸ ì™„ë£Œ: {new_filename}")
            return output_path
            
        except Exception as e:
            import traceback
            print(f" í…œí”Œë¦¿ ì—…ë°ì´íŠ¸ ì‹¤íŒ¨: {e}")
            print(f" ìƒì„¸ ì˜¤ë¥˜: {traceback.format_exc()}")
            return None
    
    def update_sheet_dates(self, workbook, year_month):
        """ì‹œíŠ¸ëª…ê³¼ í…ìŠ¤íŠ¸ì˜ ë‚ ì§œë¥¼ ìˆ˜ì§‘í•œ ë…„/ì›”ë¡œ ì—…ë°ì´íŠ¸"""
        try:
            import re
            
            # 1. ì‹œíŠ¸ëª… ì—…ë°ì´íŠ¸ (ì˜ˆ: "2025ë…„ 6ì›”" â†’ "2025ë…„ 5ì›”")
            for sheet in workbook.worksheets:
                if re.match(r'\d{4}ë…„ \d{1,2}ì›”', sheet.title):
                    old_title = sheet.title
                    sheet.title = year_month
                    print(f"   ì‹œíŠ¸ëª… ë³€ê²½: {old_title} â†’ {year_month}")
                    break
            
            # 2. B1-E1 ë³‘í•© ì…€ í…ìŠ¤íŠ¸ ì—…ë°ì´íŠ¸
            target_sheet = None
            for sheet in workbook.worksheets:
                if sheet.title == year_month:
                    target_sheet = sheet
                    break
            
            if target_sheet:
                # B1 ì…€ì˜ í…ìŠ¤íŠ¸ í™•ì¸ ë° ì—…ë°ì´íŠ¸
                b1_cell = target_sheet.cell(row=1, column=2)
                if b1_cell.value and isinstance(b1_cell.value, str):
                    old_text = b1_cell.value
                    # ë‚ ì§œ íŒ¨í„´ ì°¾ì•„ì„œ êµì²´ (ì˜ˆ: "2025ë…„ 6ì›”" â†’ "2025ë…„ 5ì›”")
                    new_text = re.sub(r'\d{4}ë…„ \d{1,2}ì›”', year_month, old_text)
                    if new_text != old_text:
                        b1_cell.value = new_text
                        print(f"   B1 ì…€ í…ìŠ¤íŠ¸ ë³€ê²½: {old_text} â†’ {new_text}")
            
            # 3. ëŒ€ì™¸ê³µë¬¸ ì‹œíŠ¸ì˜ í…ìŠ¤íŠ¸ ë° ìˆ˜ì‹ ì—…ë°ì´íŠ¸
            if 'ëŒ€ì™¸ê³µë¬¸' in [sheet.title for sheet in workbook.worksheets]:
                doc_sheet = workbook['ëŒ€ì™¸ê³µë¬¸']
                
                # B13 ì…€ ì—…ë°ì´íŠ¸
                b13_cell = doc_sheet.cell(row=13, column=2)
                if b13_cell.value and isinstance(b13_cell.value, str):
                    old_text = b13_cell.value
                    new_text = re.sub(r'\d{4}ë…„ \d{1,2}ì›”', year_month, old_text)
                    if new_text != old_text:
                        b13_cell.value = new_text
                        print(f"   ëŒ€ì™¸ê³µë¬¸ B13 ì…€ ë³€ê²½: {old_text} â†’ {new_text}")
                
                # B16 ì…€ ì—…ë°ì´íŠ¸ (B,C,D,E,F,G 16í–‰ ë³‘í•©)
                b16_cell = doc_sheet.cell(row=16, column=2)
                if b16_cell.value and isinstance(b16_cell.value, str):
                    old_text = b16_cell.value
                    new_text = re.sub(r'2025ë…„\d{1,2}ì›”', year_month, old_text)
                    if new_text == old_text:
                        new_text = re.sub(r'\d{4}ë…„\d{1,2}ì›”', year_month, old_text)
                    if new_text != old_text:
                        b16_cell.value = new_text
                        print(f"   ëŒ€ì™¸ê³µë¬¸ B16 ì…€ ë³€ê²½: {old_text} â†’ {new_text}")
                
                # ìˆ˜ì‹ ì—…ë°ì´íŠ¸ (ì‹œíŠ¸ëª… ì°¸ì¡° ë³€ê²½)
                self.update_formula_references(doc_sheet, year_month)
                        
        except Exception as e:
            print(f" ì‹œíŠ¸ ë‚ ì§œ ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}")
    
    def update_formula_references(self, doc_sheet, year_month):
        """ëŒ€ì™¸ê³µë¬¸ ì‹œíŠ¸ì˜ ìˆ˜ì‹ì—ì„œ ì‹œíŠ¸ëª… ì°¸ì¡° ì—…ë°ì´íŠ¸"""
        try:
            import re
            
            # ìˆ˜ì‹ì´ ìˆì„ ìˆ˜ ìˆëŠ” ì…€ë“¤ í™•ì¸ (24í–‰, 25í–‰ ë“±)
            cells_to_check = [
                (24, 2), (24, 3), (24, 4),  # B24, C24, D24
                (25, 2), (25, 3), (25, 4),  # B25, C25, D25
                (26, 2), (26, 3), (26, 4),  # B26, C26, D26
                (27, 2), (27, 3), (27, 4),  # B27, C27, D27
            ]
            
            for row, col in cells_to_check:
                cell = doc_sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    old_formula = cell.value
                    # ìˆ˜ì‹ì—ì„œ ì‹œíŠ¸ëª… ì°¸ì¡° íŒ¨í„´ ì°¾ì•„ì„œ êµì²´
                    # ì˜ˆ: ='2025ë…„ 6ì›”'!C4 â†’ ='2025ë…„ 5ì›”'!C4
                    new_formula = re.sub(r"'(\d{4}ë…„ \d{1,2}ì›”)'!", f"'{year_month}'!", old_formula)
                    if new_formula != old_formula:
                        cell.value = new_formula
                        print(f"    ëŒ€ì™¸ê³µë¬¸ {chr(64+col)}{row} ì…€ ìˆ˜ì‹ ë³€ê²½: {old_formula} â†’ {new_formula}")
                        
        except Exception as e:
            print(f" ìˆ˜ì‹ ì°¸ì¡° ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}")
    
    def process_sms_data(self, workbook, template_team, csv_files):
        """SMS ë°ì´í„° ì²˜ë¦¬ ë° ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ ì—…ë°ì´íŠ¸"""
        try:
            # SMS CSV íŒŒì¼ ì°¾ê¸° (ë³€í™˜ëœ CSV íŒŒì¼ë“¤ ì¤‘ì—ì„œ)
            sms_file = None
            for csv_file in csv_files:
                if 'SMS' in csv_file and csv_file.endswith('.csv'):
                    sms_file = csv_file
                    break
            
            if not sms_file:
                print("    SMS CSV íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
                print(f"   ì‚¬ìš© ê°€ëŠ¥í•œ íŒŒì¼: {csv_files}")
                return
            
            print(f"    SMS íŒŒì¼ ë°œê²¬: {sms_file}")
            
            # SMS ë°ì´í„° ì½ê¸° (CSV íŒŒì¼)
            sms_df = pd.read_csv(sms_file)
            print(f"    SMS ë°ì´í„° ë¡œë“œ: {len(sms_df)}í–‰")
            print(f"    ì»¬ëŸ¼: {list(sms_df.columns)}")
            
            # ë°œì†¡ìƒíƒœê°€ "ì„±ê³µ(ì „ë‹¬)"ì¸ ê²ƒë§Œ í•„í„°ë§
            success_df = sms_df[sms_df['ë°œì†¡ìƒíƒœ'] == 'ì„±ê³µ(ì „ë‹¬)'].copy()
            print(f"   ì„±ê³µ ì „ë‹¬ ê±´ìˆ˜: {len(success_df)}ê±´")
            
            # ë°œì‹ ë²ˆí˜¸ë¥¼ ë¬¸ìì—´ë¡œ ë³€í™˜í•˜ê³  ì •ë¦¬
            success_df['ë°œì‹ ë²ˆí˜¸_ì •ë¦¬'] = success_df['ë°œì‹ ë²ˆí˜¸'].astype(str).str.replace('.0', '')
            
            # í…œí”Œë¦¿ë³„ ë°œì‹ ë²ˆí˜¸ ë§¤í•‘ (í•˜ì´í”ˆ ì œê±°ëœ í˜•íƒœë¡œ ë¹„êµ)
            sender_mapping = {
                "CSíŒ€": "15888298",
                "ì—”í•˜ìš°ìŠ¤": "15884611", 
                "ì‚¬ì—…ì§€ì›íŒ€": "15880656"
            }
            
            print(f"    ë°œì‹ ë²ˆí˜¸ ìƒ˜í”Œ: {success_df['ë°œì‹ ë²ˆí˜¸_ì •ë¦¬'].head(10).tolist()}")
            
            # í•´ë‹¹ í…œí”Œë¦¿ì˜ ë°œì‹ ë²ˆí˜¸ë¡œ í•„í„°ë§
            target_sender = sender_mapping.get(template_team)
            if template_team == "ì‚¬ì—…ì§€ì›íŒ€":
                # ì‚¬ì—…ì§€ì›íŒ€ì˜ ê²½ìš° 15880656 ë˜ëŠ” NULL/ë¹ˆê°’ ëª¨ë‘ í¬í•¨
                team_sms_df = success_df[
                    (success_df['ë°œì‹ ë²ˆí˜¸_ì •ë¦¬'] == '15880656') | 
                    (success_df['ë°œì‹ ë²ˆí˜¸'].isna()) | 
                    (success_df['ë°œì‹ ë²ˆí˜¸'] == '') |
                    (success_df['ë°œì‹ ë²ˆí˜¸_ì •ë¦¬'] == 'nan')
                ]
            elif target_sender:
                team_sms_df = success_df[success_df['ë°œì‹ ë²ˆí˜¸_ì •ë¦¬'] == target_sender]
            else:
                team_sms_df = success_df.iloc[0:0]  # ë¹ˆ DataFrame
            
            print(f"    {template_team} í•´ë‹¹ SMS ê±´ìˆ˜: {len(team_sms_df)}ê±´")
            
            # ë¬¸ììœ í˜•ë³„ ì¹´ìš´íŠ¸
            counts = {"SMS": 0, "LMS": 0, "TALK": 0}
            
            print(f"    ë¬¸ììœ í˜• ìƒ˜í”Œ: {team_sms_df['ë¬¸ììœ í˜•'].value_counts().head()}")
            
            for _, row in team_sms_df.iterrows():
                msg_type = row.get('ë¬¸ììœ í˜•', '')
                
                if msg_type == "SMS":
                    counts["SMS"] += 1
                elif msg_type == "LMS/MMS":
                    counts["LMS"] += 1
                elif msg_type == "TALK(ì•Œë¦¼í†¡)":
                    counts["TALK"] += 1
            
            print(f"    SMS: {counts['SMS']}ê±´, LMS/MMS: {counts['LMS']}ê±´, TALK: {counts['TALK']}ê±´")
            
            # ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ ì—…ë°ì´íŠ¸
            self.update_detail_sheet(workbook, counts)
            
        except Exception as e:
            print(f"SMS ë°ì´í„° ì²˜ë¦¬ ì˜¤ë¥˜: {e}")
    
    def update_detail_sheet(self, workbook, counts):
        """ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ì˜ D13-D16 ì…€ ì—…ë°ì´íŠ¸"""
        try:
            if 'ì„¸ë¶€ë‚´ì—­' in [sheet.title for sheet in workbook.worksheets]:
                detail_sheet = workbook['ì„¸ë¶€ë‚´ì—­']
                
                # D13-D16 ì…€ì— ê±´ìˆ˜ ì…ë ¥
                detail_sheet.cell(row=13, column=4).value = counts["SMS"]    # D13: SMS ê±´ìˆ˜
                detail_sheet.cell(row=14, column=4).value = counts["LMS"]    # D14: LMS/MMS ê±´ìˆ˜
                detail_sheet.cell(row=15, column=4).value = 0                # D15: 0ìœ¼ë¡œ ì„¤ì •
                detail_sheet.cell(row=16, column=4).value = counts["TALK"]   # D16: ì•Œë¦¼í†¡ ê±´ìˆ˜
                
                print(f"    ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ ì—…ë°ì´íŠ¸ ì™„ë£Œ - D13:{counts['SMS']}, D14:{counts['LMS']}, D15:0, D16:{counts['TALK']}")
            else:
                print("    ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
                
        except Exception as e:
            print(f" ì„¸ë¶€ë‚´ì—­ ì‹œíŠ¸ ì—…ë°ì´íŠ¸ ì˜¤ë¥˜: {e}")
    
    def process_anhous_data(self, collection_date):
        """ì•¤í•˜ìš°ìŠ¤ ë°ì´í„° ì „ì²˜ë¦¬ ë©”ì¸ í•¨ìˆ˜"""
        try:
            print(" ì•¤í•˜ìš°ìŠ¤ ë°ì´í„° ì „ì²˜ë¦¬ ì‹œì‘")
            
            # 1. CSV ë³€í™˜
            csv_files = self.convert_all_collected_files()
            print(f" ë³€í™˜ ì™„ë£Œ: {len(csv_files)}ê°œ íŒŒì¼")
            
            # 2. í…œí”Œë¦¿ ì°¾ê¸°
            templates = self.find_anhous_templates()
            if not templates:
                print(" ì•¤í•˜ìš°ìŠ¤ í…œí”Œë¦¿ íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
                return False
            
            # 3. íŒ€ë³„ ë°ì´í„° ë¶„ë¥˜
            team_data = self.get_call_data_by_team(csv_files, collection_date)
            if not team_data:
                print(" íŒ€ë³„ ë°ì´í„°ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
                return False
            
            # 4. í…œí”Œë¦¿ ì—…ë°ì´íŠ¸
            print("  í…œí”Œë¦¿ íŒŒì¼ ì—…ë°ì´íŠ¸ ì¤‘...")
            print(f"   ë°œê²¬ëœ í…œí”Œë¦¿: {templates}")
            print(f"   ì‚¬ìš© ê°€ëŠ¥í•œ íŒ€: {list(team_data.keys())}")
            updated_files = []
            
            for template_name in templates:
                print(f" í…œí”Œë¦¿ ì²˜ë¦¬ ì‹œì‘: {template_name}")
                # íŒ€ ë§¤ì¹­
                template_team = None
                if 'annhouse_CS' in template_name:
                    template_team = "CSíŒ€"
                elif 'annhouse_TS' in template_name:
                    template_team = "ì—”í•˜ìš°ìŠ¤"
                elif template_name == 'annhouse.xlsx':
                    template_team = "ì‚¬ì—…ì§€ì›íŒ€"
                
                print(f"   ë§¤ì¹­ëœ íŒ€: {template_team}")
                print(f"   ì‚¬ìš© ê°€ëŠ¥í•œ íŒ€: {list(team_data.keys())}")
                
                if template_team and template_team in team_data:
                    print(f" {template_name} í…œí”Œë¦¿ ì²˜ë¦¬ ì‹œì‘ (íŒ€: {template_team})")
                    # í…œí”Œë¦¿ ë‹¤ìš´ë¡œë“œ ë° ì—…ë°ì´íŠ¸
                    template_path = self.download_template(template_name)
                    if template_path:
                        print(f" í…œí”Œë¦¿ ë‹¤ìš´ë¡œë“œ ì™„ë£Œ: {template_path}")
                        updated_path = self.update_template_with_team_data(
                            template_path, 
                            {template_team: team_data[template_team]}, 
                            collection_date,
                            template_team,
                            csv_files
                        )
                        if updated_path:
                            updated_files.append(updated_path)
                            print(f" í…œí”Œë¦¿ ì—…ë°ì´íŠ¸ ì™„ë£Œ: {updated_path}")
                    else:
                        print(f" í…œí”Œë¦¿ ë‹¤ìš´ë¡œë“œ ì‹¤íŒ¨: {template_name}")
                else:
                    print(f" í…œí”Œë¦¿ ë§¤ì¹­ ì‹¤íŒ¨: {template_name} (íŒ€: {template_team})")
            
            print(f" ì „ì²˜ë¦¬ ì™„ë£Œ: {len(updated_files)}ê°œ íŒŒì¼ ì—…ë°ì´íŠ¸")
            
            # self.cleanup_temp_folder()
            
            return True
            
        except Exception as e:
            print(f" ì „ì²˜ë¦¬ ì‹¤íŒ¨: {e}")
            return False
    
    def cleanup_temp_folder(self):
        """temp_processing í´ë” ì •ë¦¬"""
        try:
            temp_dir = "temp_processing"
            if os.path.exists(temp_dir):
                # ëª¨ë“  íŒŒì¼ ì‚­ì œ
                for filename in os.listdir(temp_dir):
                    file_path = os.path.join(temp_dir, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                        print(f"ì‚­ì œ: {filename}")
                
                print(" temp_processing í´ë” ì •ë¦¬ ì™„ë£Œ")
            else:
                print(" temp_processing í´ë”ê°€ ì¡´ì¬í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤")
        except Exception as e:
            print(f" í´ë” ì •ë¦¬ ì‹¤íŒ¨: {e}")