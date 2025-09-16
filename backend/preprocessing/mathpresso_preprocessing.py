import os
import shutil
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import re
import firebase_admin
from firebase_admin import credentials, storage
from backend.utils.secrets_manager import get_firebase_secret

class MathpressoPreprocessor:
    def __init__(self):
        self.download_dir = str(Path.home() / "Downloads")
        self.setup_firebase()
    
    def setup_firebase(self):
        """Firebase Storage 연결 설정"""
        try:
            from dotenv import load_dotenv
            
            load_dotenv()
            cred_dict = get_firebase_secret()
            
            BUCKET_NAME = os.getenv("STORAGE_BUCKET", "services-e42af.firebasestorage.app")
            
            # 기존 앱이 있으면 삭제하고 새로 초기화
            if firebase_admin._apps:
                firebase_admin.delete_app(firebase_admin.get_app())
            
            firebase_admin.initialize_app(
                credentials.Certificate(cred_dict),
                {"storageBucket": BUCKET_NAME}
            )
            
            self.bucket = storage.bucket()
            print("Firebase Storage 연결 완료")
        except Exception as e:
            print(f"Firebase 연결 실패: {e}")
            self.bucket = None
    
    def download_mathpresso_template(self):
        """Firebase에서 mathpresso.xlsx 템플릿 다운로드"""
        if not self.bucket:
            print("Firebase 연결이 없습니다")
            return None
        
        try:
            blob = self.bucket.blob("mathpresso.xlsx")
            temp_dir = os.path.join(os.getcwd(), "temp_processing")
            os.makedirs(temp_dir, exist_ok=True)
            
            local_path = os.path.join(temp_dir, "mathpresso.xlsx")
            blob.download_to_filename(local_path)
            print(f"mathpresso.xlsx 템플릿 다운로드 완료: {local_path}")
            return local_path
        except Exception as e:
            print(f"mathpresso.xlsx 템플릿 다운로드 실패: {e}")
            return None
    
    def get_bill_amount(self, company_name="매스프레소(콴다)"):
        """고지서에서 업데이트된 금액 조회"""
        try:
            import requests
            response = requests.get('http://localhost:5001/api/bill-amounts')
            if response.ok:
                bill_data = response.json()
                if company_name in bill_data:
                    amount_str = bill_data[company_name].get('amount', '')
                    if amount_str:
                        amount_clean = amount_str.replace(',', '').replace('원', '').strip()
                        if amount_clean.isdigit():
                            amount = float(amount_clean)
                            print(f"{company_name} 고지서 금액 조회: {amount:,.0f}원")
                            return amount
                        else:
                            print(f"금액 형식 오류: {amount_str}")
                            return None
                else:
                    print(f"{company_name} 고지서 정보를 찾을 수 없습니다")
                    return None
            else:
                print("고지서 정보 조회 실패")
                return None
        except Exception as e:
            print(f"고지서 금액 조회 실패: {e}")
            return None
    
    def calculate_amount_without_vat(self, total_amount):
        """부가세 10%를 제외한 금액 계산 (1원 오차 보정 포함)"""
        try:
            # 부가세 포함 금액에서 부가세 제외
            # 총금액 = 공급가액 + 부가세(공급가액의 10%)
            # 총금액 = 공급가액 × 1.1
            # 공급가액 = 총금액 / 1.1
            amount_without_vat = round(total_amount / 1.1)
            
            # 검증: 부가세 제외 금액 + 부가세 = 실제 고지서 금액인지 확인
            calculated_total = round(amount_without_vat * 1.1)
            difference = total_amount - calculated_total
            
            if difference != 0:
                # 1원 차이가 있으면 부가세 제외 금액을 보정
                amount_without_vat += difference
                print(f"1원 오차 보정: {difference:+d}원 조정")
                
            # 최종 검증
            final_total = round(amount_without_vat * 1.1)
            print(f"부가세 제외 계산: {total_amount:,.0f}원 → {amount_without_vat}원 (검증: {final_total:,}원)")
            
            return amount_without_vat
        except Exception as e:
            print(f"부가세 제외 계산 실패: {e}")
            return None
    
    def find_uploaded_file(self, company_name):
        """업로드된 파일 찾기 - temp 폴더에서 매스프레소(콴다) 파일 찾기"""
        try:
            temp_dir = os.path.join(os.getcwd(), "temp_processing")
            if os.path.exists(temp_dir):
                all_files = os.listdir(temp_dir)
                mathpresso_files = []
                
                for filename in all_files:
                    if ("매스프레소(콴다)" in filename and filename.endswith('.xlsx')):
                        file_path = os.path.join(temp_dir, filename)
                        import time
                        current_time = time.time()
                        if os.path.exists(file_path) and (current_time - os.path.getctime(file_path)) < 3600:
                            mathpresso_files.append((filename, os.path.getctime(file_path)))
                
                if mathpresso_files:
                    latest_file = sorted(mathpresso_files, key=lambda x: x[1], reverse=True)[0]
                    filename = latest_file[0]
                    file_path = os.path.join(temp_dir, filename)
                    print(f"매스프레소(콴다) 파일 발견: {filename}")
                    return file_path
                else:
                    print(f"temp_processing 폴더에서 매스프레소(콴다) 파일을 찾을 수 없습니다")
            else:
                print(f"temp_processing 폴더가 존재하지 않습니다: {temp_dir}")
            
            return None
        except Exception as e:
            print(f" 업로드된 파일 찾기 실패: {e}")
            return None
    
    def convert_xlsx_to_csv_and_count_success(self, xlsx_file_path):
        """XLSX 파일을 CSV로 변환하고 문자유형별 성공 건수 카운트"""
        try:
            # XLSX 파일 읽기
            df = pd.read_excel(xlsx_file_path)
            print(f" XLSX 파일 읽기 완료: {len(df)}행")
            
            # 필요한 컬럼 확인
            if '발송상태' not in df.columns:
                print("[발송상태] 컬럼을 찾을 수 없습니다")
                return {}
            
            if '문자유형' not in df.columns:
                print("[문자유형] 컬럼을 찾을 수 없습니다")
                return {}
            
            # 문자유형별 성공(전달) 건수 카운트
            success_df = df[df['발송상태'] == '성공(전달)']
            
            sms_count = len(success_df[success_df['문자유형'] == 'SMS'])
            lms_count = len(success_df[success_df['문자유형'] == 'LMS'])
            mms_count = len(success_df[success_df['문자유형'] == 'MMS'])
            talk_count = len(success_df[success_df['문자유형'] == 'TALK'])
            
            print(f"문자유형별 성공(전달) 건수:")
            print(f"   - SMS: {sms_count}건")
            print(f"   - LMS: {lms_count}건")
            print(f"   - MMS: {mms_count}건")
            print(f"   - TALK: {talk_count}건")
            print(f"   - 전체 행수: {len(df)}행")
            
            return {
                'sms_count': sms_count,
                'lms_count': lms_count,
                'mms_count': mms_count,
                'talk_count': talk_count
            }
                
        except Exception as e:
            print(f"XLSX 파일 처리 실패: {e}")
            return {}
    
    
    def update_mathpresso_template(self, template_path, message_counts, amount_without_vat, collection_date):
        """mathpresso.xlsx 템플릿 파일 업데이트"""
        try:
            date_obj = datetime.strptime(collection_date, '%Y-%m-%d')
            year_month = f"{date_obj.year}년 {date_obj.month:02d}월"
            date_prefix = f"{str(date_obj.year)[2:]}{date_obj.month:02d}"
            
            # 출력 파일명 생성
            output_filename = f"{date_prefix}_매스프레소(콴다)_청구내역서.xlsx"
            output_path = os.path.join(self.download_dir, output_filename)
            
            # 템플릿 파일 복사
            shutil.copy2(template_path, output_path)
            print(f"템플릿 파일 복사 완료: {output_filename}")
            
            # 워크북 로드
            workbook = load_workbook(output_path)
            
            # B9 셀에 문서번호 설정 (MMP-{년월} 형식)
            document_number = f"MMP-{date_prefix}"
            for sheet in workbook.worksheets:
                if '세부내역' in sheet.title or '대외공문' in sheet.title:
                    sheet.cell(row=9, column=2).value = f"문서번호  : {document_number}"
                    print(f"{sheet.title} B9 셀에 문서번호 설정 완료: {document_number}")
            
            # 1. 대외공문 시트 업데이트 (SK일렉링크와 동일한 로직)
            if '대외공문' in workbook.sheetnames:
                doc_sheet = workbook['대외공문']
                
                # B13 셀 업데이트
                b13_cell = doc_sheet.cell(row=13, column=2)
                if b13_cell.value and isinstance(b13_cell.value, str):
                    old_text = b13_cell.value
                    new_text = re.sub(r'\d{4}년 \d{1,2}월', year_month, old_text)
                    b13_cell.value = new_text
                    print(f"대외공문 B13 셀 업데이트: {old_text} → {new_text}")
                
                # B16 셀 업데이트 (B,C,D,E,F,G 16행 병합)
                b16_cell = doc_sheet.cell(row=16, column=2)
                if b16_cell.value and isinstance(b16_cell.value, str):
                    old_text = b16_cell.value
                    new_text = re.sub(r'2025년\d{1,2}월', year_month, old_text)
                    if new_text == old_text:
                        new_text = re.sub(r'\d{4}년\d{1,2}월', year_month, old_text)
                        if new_text == old_text:
                            new_text = re.sub(r'\d{4}년\s*\d{1,2}월', year_month, old_text)
                    b16_cell.value = new_text
                    print(f"대외공문 B16 셀 업데이트: {old_text} → {new_text}")
                
                # 하단 테이블 수식 업데이트 (B24, D24, D25)
                self.update_formula_references(doc_sheet, year_month)
            
            # 2. 시트명 변경 및 해당 시트의 B,C,D,E1 병합 셀 텍스트 업데이트
            for sheet in workbook.worksheets:
                if re.match(r'\d{4}년 \d{1,2}월', sheet.title):
                    old_title = sheet.title
                    sheet.title = year_month
                    print(f"시트명 변경: {old_title} → {year_month}")
                    
                    # B,C,D,E1 병합 셀의 텍스트 업데이트
                    b1_cell = sheet.cell(row=1, column=2)  # B1 셀
                    if b1_cell.value and isinstance(b1_cell.value, str):
                        old_text = b1_cell.value
                        new_text = re.sub(r'\d{4}년 \d{1,2}월', year_month, old_text)
                        if new_text != old_text:
                            b1_cell.value = new_text
                            print(f"{year_month} 시트 B1 셀 텍스트 업데이트: {old_text} → {new_text}")
                    break
            
            # 3. 세부내역 시트 업데이트
            if '세부내역' in workbook.sheetnames:
                detail_sheet = workbook['세부내역']
                
                # 문자유형별 성공 건수 입력
                # E12: SMS, E13: LMS, E14: MMS, E15: TALK
                if 'sms_count' in message_counts:
                    detail_sheet.cell(row=12, column=5).value = message_counts['sms_count']
                    print(f"세부내역 시트 E12 셀 업데이트: {message_counts['sms_count']}건 (SMS)")
                
                if 'lms_count' in message_counts:
                    detail_sheet.cell(row=13, column=5).value = message_counts['lms_count']
                    print(f"세부내역 시트 E13 셀 업데이트: {message_counts['lms_count']}건 (LMS)")
                
                if 'mms_count' in message_counts:
                    detail_sheet.cell(row=14, column=5).value = message_counts['mms_count']
                    print(f"세부내역 시트 E14 셀 업데이트: {message_counts['mms_count']}건 (MMS)")
                
                if 'talk_count' in message_counts:
                    detail_sheet.cell(row=15, column=5).value = message_counts['talk_count']
                    print(f"세부내역 시트 E15 셀 업데이트: {message_counts['talk_count']}건 (TALK)")
                
                # F4 셀에 부가세 제외 금액 입력 (숫자만)
                detail_sheet.cell(row=4, column=6).value = amount_without_vat
                print(f"세부내역 시트 F4 셀 업데이트: {amount_without_vat}원 (쉼표 없이)")
            
            # 파일 저장
            workbook.save(output_path)
            workbook.close()
            
            print(f"mathpresso.xlsx 템플릿 업데이트 완료: {output_filename}")
            return output_path
            
        except Exception as e:
            print(f"mathpresso.xlsx 템플릿 업데이트 실패: {e}")
            return None
    
    def update_formula_references(self, doc_sheet, year_month):
        """대외공문 시트의 수식에서 시트명 참조 업데이트"""
        try:
            # 수식이 있을 수 있는 셀들 확인 (24행, 25행 등)
            cells_to_check = [
                (24, 2), (24, 3), (24, 4),  # B24, C24, D24
                (25, 2), (25, 3), (25, 4),  # B25, C25, D25
            ]
            
            for row, col in cells_to_check:
                cell = doc_sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    old_formula = cell.value
                    # 수식에서 시트명 참조 패턴 찾아서 교체
                    new_formula = re.sub(r"'(\d{4}년 \d{1,2}월)'!", f"'{year_month}'!", old_formula)
                    if new_formula != old_formula:
                        cell.value = new_formula
                        print(f"대외공문 {chr(64+col)}{row} 셀 수식 업데이트: {old_formula} → {new_formula}")
                        
        except Exception as e:
            print(f"수식 참조 업데이트 오류: {e}")
    
    def process_mathpresso_data(self, collection_date):
        """매스프레소(콴다) 데이터 전처리 메인 함수"""
        try:
            print("매스프레소(콴다) 데이터 전처리 시작")
            
            # 1. 고지서 금액 조회
            total_amount = self.get_bill_amount("매스프레소(콴다)")
            if total_amount is None:
                print(" 매스프레소(콴다) 고지서 금액을 찾을 수 없습니다")
                return False
            
            # 2. 부가세 제외 금액 계산
            amount_without_vat = self.calculate_amount_without_vat(total_amount)
            if amount_without_vat is None:
                print(" 부가세 제외 계산 실패")
                return False
            
            # 3. 업로드된 파일 찾기
            uploaded_file_path = self.find_uploaded_file("매스프레소(콴다)")
            if uploaded_file_path is None:
                print(" 업로드된 파일을 찾을 수 없습니다")
                return False
            
            # 4. XLSX 파일을 CSV로 변환하고 문자유형별 성공 건수 카운트
            message_counts = self.convert_xlsx_to_csv_and_count_success(uploaded_file_path)
            if not message_counts:
                print("문자유형별 건수 카운트 실패")
                return False
            print(f"문자유형별 성공(전달) 건수: {message_counts}")
            
            # 5. mathpresso.xlsx 템플릿 다운로드
            template_path = self.download_mathpresso_template()
            if template_path is None:
                print("mathpresso.xlsx 템플릿 다운로드 실패")
                return False
            
            # 6. 템플릿 업데이트 및 청구서 생성
            final_invoice_path = self.update_mathpresso_template(template_path, message_counts, amount_without_vat, collection_date)
            if final_invoice_path is None:
                print("매스프레소(콴다) 청구서 생성 실패")
                return False
            
            print(f"매스프레소(콴다) 전처리 완료! 파일 생성:")
            print(f"   - 매스프레소(콴다) 청구내역서: {os.path.basename(final_invoice_path)}")
            
            # temp_processing 폴더 정리
            self.cleanup_temp_folder()
            
            return True
            
        except Exception as e:
            import traceback
            print(f"매스프레소(콴다) 전처리 실패: {e}")
            print(f"상세 에러: {traceback.format_exc()}")
            return False

    def cleanup_temp_folder(self):
        """temp_processing 폴더 정리"""
        try:
            temp_dir = "temp_processing"
            if os.path.exists(temp_dir):
                # 모든 파일 삭제
                for filename in os.listdir(temp_dir):
                    file_path = os.path.join(temp_dir, filename)
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                        print(f"삭제: {filename}")
                
                print("temp_processing 폴더 정리 완료")
            else:
                print("temp_processing 폴더가 존재하지 않습니다")
        except Exception as e:
            print(f"폴더 정리 실패: {e}")
